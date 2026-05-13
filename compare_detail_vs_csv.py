"""订单明细 vs 3 个 CSV 对账（仅 WT_ 订单号）"""
import sys, os, csv
from collections import Counter, defaultdict
from datetime import datetime, date as date_t
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding='utf-8')

CSV_FILES = [
    r'D:\snowmeet\snowmeet_ai_doc\ZuLinDingDan_2025-10-16_2025-11-30.csv',
    r'D:\snowmeet\snowmeet_ai_doc\ZuLinDingDan_2025-12-01_2026-01-31.csv',
    r'D:\snowmeet\snowmeet_ai_doc\ZuLinDingDan_2026-02-01_2026-04-15.csv',
]
EXCEL = r'D:\snowmeet\snowmeet_ai_doc\wanlong_rent_orders_2025-10-15_2026-04-15.xlsx'
OUT = r'D:\snowmeet\snowmeet_ai_doc\comparison_report.xlsx'


def f2(v):
    try: return round(float(v or 0), 2)
    except: return 0.0


def d2s(v):
    if v is None: return ''
    if isinstance(v, (datetime, date_t)): return v.strftime('%Y-%m-%d')
    return str(v)


# ---- 加载 CSV ----
csv_rows = []
for f in CSV_FILES:
    with open(f, encoding='utf-8') as cf:
        for r in csv.DictReader(cf):
            code = (r.get('订单号') or '').strip()
            if not code.startswith('WT_'):
                continue
            csv_rows.append({
                '订单号': code,
                '订单类型': (r.get('订单类型') or '').strip(),
                '商品名称': (r.get('商品名称') or '').strip(),
                '订单日期': (r.get('订单日期') or '').strip(),
                '订单时间': (r.get('订单时间') or '').strip(),
                '起租日期': (r.get('起租日期') or '').strip(),
                '退租日期': (r.get('退租日期') or '').strip(),
                '结算日期': (r.get('结算日期') or '').strip(),
                '租金': f2(r.get('租金')),
                '_来源': os.path.basename(f),
            })

# ---- 加载 Excel ----
wb_in = load_workbook(EXCEL, read_only=True, data_only=True)
ws_in = wb_in['订单明细']
hdr = next(ws_in.iter_rows(max_row=1, values_only=True))
xl_rows = []
for r in ws_in.iter_rows(min_row=2, values_only=True):
    d = dict(zip(hdr, r))
    xl_rows.append({
        '订单号': (d['订单号'] or '').strip(),
        '商品名称': (d['租赁商品名称'] or '').strip(),
        '起租日期': d2s(d['起租日期']),
        '退租日期': d2s(d['退租日期']),
        '租金': f2(d['租金总额']),
        '超时费': f2(d['超时费']),
        '损坏赔偿': f2(d['损坏赔偿']),
    })

# ---- multiset 配对 ----
# 键：(订单号, 商品名称)；不同金额可能存在，所以租金不进 key（金额差异另算）
csv_by_key = defaultdict(list)
for r in csv_rows: csv_by_key[(r['订单号'], r['商品名称'])].append(r)
xl_by_key = defaultdict(list)
for r in xl_rows: xl_by_key[(r['订单号'], r['商品名称'])].append(r)

all_keys = set(csv_by_key) | set(xl_by_key)
matched_pairs = []      # (csv_row, xl_row) — 已配对
only_csv = []
only_xl = []
for key in all_keys:
    cl = list(csv_by_key.get(key, []))
    xl = list(xl_by_key.get(key, []))
    n = min(len(cl), len(xl))
    for i in range(n):
        matched_pairs.append((cl[i], xl[i]))
    only_csv.extend(cl[n:])
    only_xl.extend(xl[n:])

# 金额差异（已配对但金额不同）
mismatch = [(c, x) for c, x in matched_pairs if abs(c['租金'] - x['租金']) > 0.005]

# 概览
csv_total = sum(r['租金'] for r in csv_rows)
xl_total = sum(r['租金'] for r in xl_rows)
matched_csv_sum = sum(c['租金'] for c, _ in matched_pairs)
matched_xl_sum = sum(x['租金'] for _, x in matched_pairs)
only_csv_sum = sum(r['租金'] for r in only_csv)
only_xl_sum = sum(r['租金'] for r in only_xl)

# 订单类型分布
csv_type_cnt = Counter(r['订单类型'] for r in csv_rows)

# 按订单号汇总（行数 + 金额）
order_set = set(r['订单号'] for r in csv_rows) | set(r['订单号'] for r in xl_rows)
order_stat = []
for code in sorted(order_set):
    cl = [r for r in csv_rows if r['订单号'] == code]
    xl = [r for r in xl_rows if r['订单号'] == code]
    cs, xs = sum(r['租金'] for r in cl), sum(r['租金'] for r in xl)
    cn, xn = len(cl), len(xl)
    if cn != xn or abs(cs - xs) > 0.005:
        order_stat.append({
            '订单号': code,
            'CSV笔数': cn, 'CSV租金合计': round(cs, 2),
            'Excel笔数': xn, 'Excel租金合计': round(xs, 2),
            '笔数差(Excel-CSV)': xn - cn,
            '租金差(Excel-CSV)': round(xs - cs, 2),
        })

# ---- 写报告 ----
wb = Workbook()


def write_sheet(ws, title, color, headers, rows):
    ws.title = title
    ws.append(headers)
    fill = PatternFill('solid', fgColor=color)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    for row in rows:
        ws.append(list(row.values()) if isinstance(row, dict) else list(row))
    ws.freeze_panes = 'A2'
    for c in range(1, len(headers) + 1):
        col = get_column_letter(c)
        max_len = sum(2 if ord(ch) > 127 else 1 for ch in str(headers[c - 1]))
        for r in range(2, min(len(rows), 200) + 2):
            v = ws.cell(row=r, column=c).value
            if v is None: continue
            s = str(v)
            wlen = sum(2 if ord(ch) > 127 else 1 for ch in s)
            if wlen > max_len: max_len = wlen
        ws.column_dimensions[col].width = min(max_len + 2, 40)


# Sheet 1: 概览
overview = [
    {'项目': 'CSV 行数 (WT_)', '值': len(csv_rows)},
    {'项目': '  其中 正常订单', '值': csv_type_cnt.get('正常订单', 0)},
    {'项目': '  其中 临时订单', '值': csv_type_cnt.get('临时订单', 0)},
    {'项目': 'Excel 订单明细行数', '值': len(xl_rows)},
    {'项目': '行数差 (Excel - CSV)', '值': len(xl_rows) - len(csv_rows)},
    {'项目': '——————', '值': '——————'},
    {'项目': 'CSV 租金合计', '值': round(csv_total, 2)},
    {'项目': 'Excel 租金合计', '值': round(xl_total, 2)},
    {'项目': '金额差 (Excel - CSV)', '值': round(xl_total - csv_total, 2)},
    {'项目': '——————', '值': '——————'},
    {'项目': '配对成功 行数', '值': len(matched_pairs)},
    {'项目': '  其中 金额一致', '值': len(matched_pairs) - len(mismatch)},
    {'项目': '  其中 金额不一致', '值': len(mismatch)},
    {'项目': '配对成功 CSV 金额合计', '值': round(matched_csv_sum, 2)},
    {'项目': '配对成功 Excel 金额合计', '值': round(matched_xl_sum, 2)},
    {'项目': '——————', '值': '——————'},
    {'项目': '仅 CSV 有 行数', '值': len(only_csv)},
    {'项目': '仅 CSV 有 金额合计', '值': round(only_csv_sum, 2)},
    {'项目': '仅 Excel 有 行数', '值': len(only_xl)},
    {'项目': '仅 Excel 有 金额合计', '值': round(only_xl_sum, 2)},
    {'项目': '——————', '值': '——————'},
    {'项目': '订单号差异数（笔数或金额）', '值': len(order_stat)},
]
write_sheet(wb.active, '概览', '1F4E78', ['项目', '值'], overview)

# Sheet 2: 仅 CSV 有
write_sheet(wb.create_sheet(), '仅CSV有', 'C0392B',
            ['订单号', '订单类型', '商品名称', '订单日期', '订单时间', '起租日期', '退租日期', '结算日期', '租金', '_来源'],
            sorted(only_csv, key=lambda r: (r['订单号'], r['商品名称'])))

# Sheet 3: 仅 Excel 有
write_sheet(wb.create_sheet(), '仅Excel有', '8E44AD',
            ['订单号', '商品名称', '起租日期', '退租日期', '租金', '超时费', '损坏赔偿'],
            sorted(only_xl, key=lambda r: (r['订单号'], r['商品名称'])))

# Sheet 4: 金额不一致
mis_rows = []
for c, x in mismatch:
    mis_rows.append({
        '订单号': c['订单号'],
        '商品名称': c['商品名称'],
        '订单日期': c['订单日期'],
        'CSV租金': c['租金'],
        'Excel租金': x['租金'],
        '差额(Excel-CSV)': round(x['租金'] - c['租金'], 2),
        'Excel起租日期': x['起租日期'],
        'Excel退租日期': x['退租日期'],
    })
mis_rows.sort(key=lambda r: (r['订单号'], r['商品名称']))
write_sheet(wb.create_sheet(), '金额不一致', 'D68910',
            ['订单号', '商品名称', '订单日期', 'CSV租金', 'Excel租金', '差额(Excel-CSV)', 'Excel起租日期', 'Excel退租日期'],
            mis_rows)

# Sheet 5: 按订单号差异汇总
write_sheet(wb.create_sheet(), '按订单号差异', '2E7D32',
            ['订单号', 'CSV笔数', 'CSV租金合计', 'Excel笔数', 'Excel租金合计', '笔数差(Excel-CSV)', '租金差(Excel-CSV)'],
            sorted(order_stat, key=lambda r: (-abs(r['笔数差(Excel-CSV)']), -abs(r['租金差(Excel-CSV)']))))

wb.save(OUT)

# ---- 打印汇总 ----
print(f'\n=== 对账完成: {OUT} ===\n')
for r in overview:
    print(f'  {r["项目"]:30s}  {r["值"]}')
print(f'\n=== 前 5 个差异订单（按差额绝对值排序）===')
for r in sorted(order_stat, key=lambda r: -abs(r['租金差(Excel-CSV)']))[:5]:
    print(f'  {r}')
