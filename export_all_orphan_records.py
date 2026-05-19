"""一次性产物：反向核对导出 —— all_销售单列表.xls 中不在四店任何
「年度零售明细」的孤儿记录，按归因分类，供人工逐项核。

孤儿定义：all 销售明细单1 的 单据编号，未被四店报表（南山/万龙体验/
万龙服务/崇礼）任一 年度零售明细 以「七色米订单号」匹配消费。

输出 all_销售单列表_孤儿记录.xlsx：
  sheet1 孤儿明细：归因 + 原 34 列（all 销售明细单1 原始行）
  sheet2 孤儿汇总：单据编号 / 门店 / 明细行数 / 归因类别
排序：待查类在前；待查行整行标红 FF9999，预期类不上色。
"""
import os
import sys
from collections import Counter, defaultdict

import xlrd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8")

BASE = os.path.dirname(os.path.abspath(__file__))
ALL_XLS = os.path.join(BASE, "all_销售单列表.xls")
OUT = os.path.join(BASE, "all_销售单列表_孤儿记录.xlsx")
DETAIL_SHEET = "销售明细单1"
FILES = {
    "南山": "nanshan_retail_orders_fy_2025-05-01_2026-04-30.xlsx",
    "万龙体验": "wanlong_retail_orders_fy_2025-05-01_2026-04-30.xlsx",
    "万龙服务": "wanlong_service_retail_orders_fy_2025-05-01_2026-04-30.xlsx",
    "崇礼": "chongli_retail_orders_fy_2025-05-01_2026-04-30.xlsx",
}
HEADER_FILL = "1F4E78"
RED = "FF9999"          # 待查（报表无七色米号）
MONEY_COLS_NAME = {"数量", "单价", "折扣", "折后单价", "总额", "成本额"}


def visual_len(s):
    return sum(2 if ord(ch) > 127 else 1 for ch in str(s))


def main():
    wb = xlrd.open_workbook(ALL_XLS)
    sh = wb.sheet_by_name(DETAIL_SHEET)
    hdr = [str(sh.cell_value(0, c)).strip() for c in range(sh.ncols)]
    dc = hdr.index("单据编号")
    sc = hdr.index("所属门店")
    doc_rows = defaultdict(list)
    doc_shop = {}
    for r in range(1, sh.nrows):
        d = str(sh.cell_value(r, dc)).strip()
        if not d:
            continue
        row = [sh.cell_value(r, c) for c in range(sh.ncols)]
        doc_rows[d].append(row)
        doc_shop.setdefault(d, str(sh.cell_value(r, sc)).strip())
    all_docs = set(doc_rows)
    n_all_rows = sum(len(v) for v in doc_rows.values())

    consumed = set()
    report_mi7 = {}   # 七色米号 -> 正/闭 状态（含关闭/剔除单，反查归因用）
    for nm, f in FILES.items():
        mwb = openpyxl.load_workbook(os.path.join(BASE, f), read_only=True)
        dws = mwb["年度零售明细"]
        drows = list(dws.iter_rows(values_only=True))
        dh = list(drows[0])
        ki = dh.index("七色米订单号")
        gi = dh.index("商品编号")
        last = None
        for r in drows[1:]:
            m = r[ki]
            if m is not None and str(m).strip():
                last = str(m).strip()
            if r[gi] is not None and str(r[gi]).strip() and last:
                consumed.add(last)
        sws = mwb["年度零售"]
        srows = list(sws.iter_rows(values_only=True))
        sh2 = list(srows[0])
        ski = sh2.index("七色米订单号")
        sci = sh2.index("正/闭")
        for r in srows[1:]:
            if not any(v is not None for v in r):
                continue
            m = r[ski]
            if m is not None and str(m).strip():
                report_mi7[str(m).strip()] = str(r[sci]).strip()
        mwb.close()

    orphan = all_docs - consumed

    def categorize(d):
        shop = doc_shop[d]
        if d in report_mi7:
            stt = report_mi7[d]
            return ("报表内·已关闭(删除)" if stt == "关闭"
                    else "报表内·剔除测试单(删除)")
        if shop == "【总部】":
            return "总部·无财年零售报表"
        if shop == "【崇礼-万龙店】":
            return "崇礼万龙店·无财年零售报表"
        if shop == "【崇礼-旗舰店】":
            return "崇礼旗舰·报表无七色米号(待查)"
        if shop == "【北京-南山店】":
            return "南山·报表无七色米号(待查)"
        return f"{shop}·无对应报表"

    cat_of = {d: categorize(d) for d in orphan}
    # 排序权重：待查在前
    order = {
        "崇礼旗舰·报表无七色米号(待查)": 0,
        "南山·报表无七色米号(待查)": 1,
        "报表内·已关闭(删除)": 2,
        "报表内·剔除测试单(删除)": 3,
        "崇礼万龙店·无财年零售报表": 4,
        "总部·无财年零售报表": 5,
    }

    def sort_key(d):
        return (order.get(cat_of[d], 9), cat_of[d], d)

    docs_sorted = sorted(orphan, key=sort_key)
    is_pending = {d: cat_of[d].endswith("(待查)") for d in orphan}

    owb = openpyxl.Workbook()
    # ===== sheet1 孤儿明细 =====
    ws = owb.active
    ws.title = "孤儿明细"
    out_hdr = ["归因类别"] + hdr
    ws.append(out_hdr)
    hf = PatternFill("solid", fgColor=HEADER_FILL)
    for c in range(1, len(out_hdr) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = hf
        cell.alignment = Alignment(horizontal="center", vertical="center")
    red_fill = PatternFill("solid", fgColor=RED)
    money_idx = [hdr.index(n) for n in hdr if n in MONEY_COLS_NAME]
    rr = 1
    for d in docs_sorted:
        for row in doc_rows[d]:
            ws.append([cat_of[d]] + row)
            rr += 1
            if is_pending[d]:
                for c in range(1, len(out_hdr) + 1):
                    ws.cell(row=rr, column=c).fill = red_fill
            for mi in money_idx:
                cl = ws.cell(row=rr, column=2 + mi)
                if isinstance(cl.value, (int, float)):
                    cl.number_format = "0.00"
    ws.freeze_panes = "B2"

    # ===== sheet2 孤儿汇总 =====
    ws2 = owb.create_sheet("孤儿汇总")
    sum_hdr = ["归因类别", "单据编号", "所属门店", "明细行数"]
    ws2.append(sum_hdr)
    for c in range(1, len(sum_hdr) + 1):
        cell = ws2.cell(row=1, column=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = hf
        cell.alignment = Alignment(horizontal="center", vertical="center")
    rr = 1
    for d in docs_sorted:
        ws2.append([cat_of[d], d, doc_shop[d], len(doc_rows[d])])
        rr += 1
        if is_pending[d]:
            for c in range(1, len(sum_hdr) + 1):
                ws2.cell(row=rr, column=c).fill = red_fill
    ws2.freeze_panes = "A2"

    # 列宽自适应
    for sheet, headers in ((ws, out_hdr), (ws2, sum_hdr)):
        for c in range(1, len(headers) + 1):
            mx = visual_len(headers[c - 1])
            for r in range(2, min(sheet.max_row, 300) + 1):
                v = sheet.cell(row=r, column=c).value
                if v is not None:
                    mx = max(mx, visual_len(v))
            sheet.column_dimensions[get_column_letter(c)].width = min(mx + 2, 40)

    owb.save(OUT)

    # ===== 控制台汇总 =====
    cnt = Counter(cat_of.values())
    rowcnt = Counter()
    for d in orphan:
        rowcnt[cat_of[d]] += len(doc_rows[d])
    print(f"all 销售明细单1: {len(all_docs)} 单据 / {n_all_rows} 明细行")
    print(f"四店消费: {len(all_docs & consumed)} 单据")
    print(f"孤儿: {len(orphan)} 单据 / "
          f"{sum(len(doc_rows[d]) for d in orphan)} 明细行\n")
    print("归因（单据数 / 明细行数）：")
    for k in sorted(cnt, key=lambda x: order.get(x, 9)):
        flag = "  ← 待查" if k.endswith("(待查)") else ""
        print(f"  {k:32s} {cnt[k]:4d} / {rowcnt[k]:4d}{flag}")
    print(f"\n[OK] 导出 {OUT}  ({os.path.getsize(OUT) / 1024:.1f} KB)")
    print("  sheet: 孤儿明细 (210 行级) / 孤儿汇总 (124 单据级)；待查行标红")


if __name__ == "__main__":
    main()
