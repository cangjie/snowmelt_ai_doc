"""给 {prefix}_care_orders_fy_{start}_{end}.xlsx 添加 sheet「年度养护明细」：
  关联「年度养护」（订单级） + care 表（设备级明细） + care_task 表（7 名执行人）。

明细列（15 care + 7 staff = 22）：
  装备 / 品牌 / 型号 / 尺码 / 加急 / 修刃 / 打蜡 / 刮蜡 / 维修 /
  维修费 / 养护费 / 直减 / 完成 / 取板日期 / 备注 /
  安全检查人 / 修刃人 / 机打蜡人 / 热打蜡人 / 刮蜡人 / 维修人 / 发板人

task_name → 列映射（用户确认）：
  安全检查人 ← '安全检查'
  修刃人 ← '修刃'
  机打蜡人 ← '机打蜡'
  热打蜡人 ← '热蜡' ∪ '打蜡'（历史 '打蜡' 单合并到此列）
  刮蜡人 ← '刮蜡'
  维修人 ← '维修'
  发板人 ← '发板'

每列取该 care 下符合 task_name 的所有 staff_id 关联 staff.name，
去重后用 ';' 连接（多次执行/多人执行场景兼容）。

形态：
  - 订单级列（即「年度养护」原全部 N 列）放左侧，一单多 care 纵向合并
  - 明细 22 列在右侧，每个 care 各一行
  - 一单无 care 行 → 保留单行明细列空
  - 多 care 订单（M ≥ 2）整行底色浅蓝 EAF2FB
  - 表头 1F4E78 蓝底白字粗体；freeze A2

幂等：「年度养护明细」存在则删重建；其他 sheet 不动。

用法：
  py add_care_detail_merged_sheet.py --xlsx wanlong_service_care_orders_fy_2025-05-01_2026-04-30.xlsx --shop 万龙服务中心
"""
import argparse
import os
import sys
from collections import defaultdict
from datetime import datetime, timedelta

import pyodbc
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding='utf-8')

MAIN_SHEET = '年度养护'
DETAIL_SHEET = '年度养护明细'
HEADER_COLOR = '1F4E78'
MULTI_FILL = 'EAF2FB'

CARE_COLS = ['装备', '品牌', '型号', '尺码', '加急',
             '修刃', '打蜡', '刮蜡', '维修',
             '维修费', '养护费', '直减', '完成',
             '取板日期', '备注']
STAFF_COLS = ['安全检查人', '修刃人', '机打蜡人', '热打蜡人',
              '刮蜡人', '维修人', '发板人']
DETAIL_COLS = CARE_COLS + STAFF_COLS

MONEY_DETAIL = {'维修费', '养护费', '直减'}
BOOL_DETAIL = {'加急', '修刃', '打蜡', '刮蜡', '维修', '完成'}

# task_name → 列名（多 task_name 共享同列时拼 list）
TASK_TO_COL = {
    '安全检查': '安全检查人',
    '修刃': '修刃人',
    '机打蜡': '机打蜡人',
    '热蜡': '热打蜡人',
    '打蜡': '热打蜡人',
    '刮蜡': '刮蜡人',
    '维修': '维修人',
    '发板': '发板人',
}

CONN = ('DRIVER={ODBC Driver 18 for SQL Server};SERVER=tcp:100.28.143.19,1433;'
        'DATABASE=snowmeet_new;UID=claude;PWD=abcd123!@#;'
        'Encrypt=yes;TrustServerCertificate=yes;Connection Timeout=30;')


def parse_args():
    p = argparse.ArgumentParser(
        description='给养护财年 xlsx 添加「年度养护明细」合并 sheet',
        formatter_class=argparse.RawDescriptionHelpFormatter, epilog=__doc__)
    p.add_argument('--xlsx', required=True, help='目标 xlsx 路径')
    p.add_argument('--shop', required=True, help='店铺名（DB order.shop）')
    p.add_argument('--start', default='2025-05-01', help='biz_date 起始（含），默认 2025-05-01')
    p.add_argument('--end', default='2026-04-30', help='biz_date 截止（含），默认 2026-04-30')
    return p.parse_args()


def bool_label(v):
    if v is None:
        return None
    return '是' if int(v) == 1 else '否'


def fetch_care_rows(shop, start, end_excl):
    """返回 {order_code: [care_row_dict, ...]}"""
    cn = pyodbc.connect(CONN)
    cur = cn.cursor()
    cur.execute("""SELECT o.code, c.id, c.equipment, c.brand, c.series, c.scale, c.urgent,
                          c.need_edge, c.need_wax, c.need_unwax, c.need_repair,
                          c.repair_charge, c.common_charge, c.discount, c.finish,
                          c.member_pick_date, c.memo
                   FROM care c JOIN [order] o ON o.id = c.order_id
                   WHERE o.shop = ? AND o.[type] = N'养护'
                     AND o.biz_date >= ? AND o.biz_date < ?
                     AND o.valid = 1 AND o.code IS NOT NULL AND LTRIM(RTRIM(o.code)) <> N''
                     AND c.valid = 1
                   ORDER BY o.id, c.id""", shop, start, end_excl)
    by_code = defaultdict(list)
    care_ids = []
    for r in cur.fetchall():
        code = r[0].strip() if r[0] else None
        if not code:
            continue
        care_ids.append(r[1])
        by_code[code].append({
            'care_id': r[1],
            '装备': r[2], '品牌': r[3], '型号': r[4], '尺码': r[5],
            '加急': bool_label(r[6]),
            '修刃': bool_label(r[7]),
            '打蜡': bool_label(r[8]),
            '刮蜡': bool_label(r[9]),
            '维修': bool_label(r[10]),
            '维修费': r[11], '养护费': r[12], '直减': r[13],
            '完成': bool_label(r[14]),
            '取板日期': r[15],
            '备注': r[16],
        })
    cn.close()
    return by_code, care_ids


def fetch_staff_by_care(care_ids):
    """返回 {care_id: {staff_col: 'name;name;...'}}"""
    if not care_ids:
        return {}
    cn = pyodbc.connect(CONN)
    cur = cn.cursor()
    # 用临时表传 ids（避免 IN list 长度限制）
    cur.execute('CREATE TABLE #ids (id INT PRIMARY KEY)')
    # fast_executemany 批量插
    cur.fast_executemany = True
    cur.executemany('INSERT INTO #ids (id) VALUES (?)', [(i,) for i in care_ids])
    cur.execute("""SELECT ct.care_id, ct.task_name, s.name, ct.id
                   FROM care_task ct
                   JOIN #ids tmp ON tmp.id = ct.care_id
                   LEFT JOIN staff s ON s.id = ct.staff_id
                   WHERE ct.valid = 1 AND ct.staff_id IS NOT NULL AND s.name IS NOT NULL
                   ORDER BY ct.care_id, ct.id""")
    # care_id → col_name → list of names (按 ct.id 升序，去重保序)
    by_care = defaultdict(lambda: defaultdict(list))
    for r in cur.fetchall():
        care_id, task_name, name, _ = r
        col = TASK_TO_COL.get(task_name)
        if col is None:
            continue
        if name not in by_care[care_id][col]:
            by_care[care_id][col].append(name)
    cur.execute('DROP TABLE #ids')
    cn.close()
    # 拼接
    return {cid: {col: ';'.join(names) for col, names in cols.items()}
            for cid, cols in by_care.items()}


def main():
    args = parse_args()
    xlsx = os.path.abspath(args.xlsx)
    end_excl = (datetime.strptime(args.end, '%Y-%m-%d') + timedelta(days=1)).strftime('%Y-%m-%d')

    if not os.path.exists(xlsx):
        raise SystemExit(f'xlsx 不存在: {xlsx}')

    print(f'读 SQL care 明细（{args.shop} / {args.start} ~ {end_excl}）...')
    by_code, care_ids = fetch_care_rows(args.shop, args.start, end_excl)
    total_cares = sum(len(v) for v in by_code.values())
    multi_orders = sum(1 for v in by_code.values() if len(v) > 1)
    print(f'  覆盖订单: {len(by_code)}（含 {multi_orders} 单多 care），care 总数: {total_cares}')

    print(f'读 care_task 关联员工（{len(care_ids)} 个 care_id）...')
    staff_map = fetch_staff_by_care(care_ids)
    print(f'  有员工的 care: {len(staff_map)}')

    print(f'打开 {xlsx}')
    wb = load_workbook(xlsx)
    if MAIN_SHEET not in wb.sheetnames:
        raise SystemExit(f'缺主 sheet「{MAIN_SHEET}」: {wb.sheetnames}')

    # 幂等：删旧 DETAIL_SHEET
    if DETAIL_SHEET in wb.sheetnames:
        print(f'  「{DETAIL_SHEET}」已存在，删除重建（幂等）')
        del wb[DETAIL_SHEET]

    main_ws = wb[MAIN_SHEET]
    main_headers = [main_ws.cell(row=1, column=c).value for c in range(1, main_ws.max_column + 1)]
    n_order_cols = len(main_headers)
    print(f'  年度养护订单级列数 N = {n_order_cols}')

    main_rows = []
    for r in range(2, main_ws.max_row + 1):
        row = [main_ws.cell(row=r, column=c).value for c in range(1, n_order_cols + 1)]
        main_rows.append(row)

    if '订单号' not in main_headers:
        raise SystemExit('年度养护找不到「订单号」列')
    code_idx = main_headers.index('订单号')

    # 订单级金额列（用于保留数字格式）
    money_order_cols = set()
    for ci, h in enumerate(main_headers, start=1):
        if h in ('支付合计', '退款合计', '订单结余', '维修费合计', '普通养护费合计',
                 '减免合计', '卡券减免合计', '养护直减合计', '应分账金额', '实分账金额',
                 '待分账金额', '支付总金额', '退款总金额') or (isinstance(h, str) and h.endswith('】金额')):
            money_order_cols.add(ci)

    # 新 sheet
    ws = wb.create_sheet(DETAIL_SHEET)
    header_font = Font(bold=True, color='FFFFFF', name='Calibri', size=11)
    header_fill = PatternFill('solid', fgColor=HEADER_COLOR)
    multi_fill = PatternFill('solid', fgColor=MULTI_FILL)
    center = Alignment(horizontal='center', vertical='center')

    all_headers = list(main_headers) + DETAIL_COLS
    for ci, h in enumerate(all_headers, start=1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center

    out_row = 2
    merges = []
    multi_row_ranges = []

    for main_row in main_rows:
        code = main_row[code_idx]
        key = code.strip() if isinstance(code, str) else None
        care_list = by_code.get(key, []) if key else []
        M = max(len(care_list), 1)

        # 订单级列（仅第一行写值，其余等同合并视觉）
        for ci, v in enumerate(main_row, start=1):
            ws.cell(row=out_row, column=ci, value=v)

        # 明细列
        for k in range(M):
            r = out_row + k
            if care_list and k < len(care_list):
                care = care_list[k]
                staff_cols = staff_map.get(care['care_id'], {})
                for di, col_name in enumerate(DETAIL_COLS):
                    if col_name in STAFF_COLS:
                        dv = staff_cols.get(col_name)
                    else:
                        dv = care.get(col_name)
                    c = ws.cell(row=r, column=n_order_cols + 1 + di, value=dv)
                    if col_name in MONEY_DETAIL and dv is not None:
                        c.number_format = '0.00'
            # 无 care 时所有明细列默认 None（不写）

        if M > 1:
            for ci in range(1, n_order_cols + 1):
                merges.append((ci, out_row, out_row + M - 1))
            multi_row_ranges.append((out_row, out_row + M - 1))

        out_row += M

    last_row = out_row - 1
    print(f'  写入数据行: {last_row - 1} 行（含合并展开）')

    for ci, start_r, end_r in merges:
        col_letter = get_column_letter(ci)
        ws.merge_cells(f'{col_letter}{start_r}:{col_letter}{end_r}')

    # 多 care 订单整行浅蓝
    for start_r, end_r in multi_row_ranges:
        for r in range(start_r, end_r + 1):
            for ci in range(1, ws.max_column + 1):
                cell = ws.cell(row=r, column=ci)
                if cell.fill.fill_type is None:
                    cell.fill = multi_fill

    # 订单级金额列格式
    for ci in money_order_cols:
        for r in range(2, last_row + 1):
            cell = ws.cell(row=r, column=ci)
            if cell.value is not None and isinstance(cell.value, (int, float)):
                cell.number_format = '0.00'

    ws.freeze_panes = 'A2'

    for ci in range(1, len(all_headers) + 1):
        col_letter = get_column_letter(ci)
        max_w = sum(2 if ord(ch) > 127 else 1 for ch in str(all_headers[ci - 1]))
        for r in range(2, min(last_row + 1, 200)):
            v = ws.cell(row=r, column=ci).value
            if v is None:
                continue
            s = v.strftime('%Y-%m-%d %H:%M:%S') if hasattr(v, 'strftime') else str(v)
            w = sum(2 if ord(ch) > 127 else 1 for ch in s)
            if w > max_w:
                max_w = w
        ws.column_dimensions[col_letter].width = min(max_w + 2, 36)

    print(f'保存 {xlsx}')
    wb.save(xlsx)
    print(f'  sheets: {load_workbook(xlsx, read_only=True).sheetnames}')
    print('完成')


if __name__ == '__main__':
    main()
