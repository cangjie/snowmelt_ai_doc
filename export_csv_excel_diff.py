"""仅生成两个 sheet 的差异表：仅CSV有 / 仅Excel有"""
import sys, os, csv
from collections import defaultdict
from datetime import datetime, date as date_t
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding='utf-8')

CSV_FILES = [
    r'D:\snowmeet\snowmeet_ai_doc\ZuLinDingDan_2025-10-16_2025-11-30.csv',
    r'D:\snowmeet\snowmeet_ai_doc\ZuLinDingDan_2025-12-01_2026-01-31.csv',
    r'D:\snowmeet\snowmeet_ai_doc\ZuLinDingDan_2026-02-01_2026-04-15.csv',
]
EXCEL = r'D:\snowmeet\snowmeet_ai_doc\wanlong_rent_orders_2025-10-15_2026-04-15.xlsx'
OUT = r'D:\snowmeet\snowmeet_ai_doc\csv_excel_diff.xlsx'


def f2(v):
    try: return round(float(v or 0), 2)
    except: return 0.0


def d2s(v):
    if v is None: return ''
    if isinstance(v, (datetime, date_t)): return v.strftime('%Y-%m-%d')
    return str(v)


# CSV (仅 WT_)
csv_rows = []
for f in CSV_FILES:
    with open(f, encoding='utf-8') as cf:
        for r in csv.DictReader(cf):
            code = (r.get('订单号') or '').strip()
            if not code.startswith('WT_'): continue
            csv_rows.append({
                '订单号': code,
                '订单类型': (r.get('订单类型') or '').strip(),
                '商品名称': (r.get('商品名称') or '').strip(),
                '订单日期': (r.get('订单日期') or '').strip(),
                '订单时间': (r.get('订单时间') or '').strip(),
                '起租日期': (r.get('起租日期') or '').strip(),
                '退租日期': (r.get('退租日期') or '').strip(),
                '结算日期': (r.get('结算日期') or '').strip(),
                '租金': f2(r.get('租金')),
                '_来源CSV': os.path.basename(f),
            })

# Excel
wb_in = load_workbook(EXCEL, read_only=True, data_only=True)
ws_in = wb_in['订单明细']
hdr = next(ws_in.iter_rows(max_row=1, values_only=True))
xl_rows = []
for r in ws_in.iter_rows(min_row=2, values_only=True):
    d = dict(zip(hdr, r))
    xl_rows.append({
        '订单号': (d['订单号'] or '').strip(),
        '租赁商品名称': (d['租赁商品名称'] or '').strip(),
        '起租日期': d2s(d['起租日期']),
        '起租时间': d['起租时间'] or '',
        '退租日期': d2s(d['退租日期']),
        '退租时间': d['退租时间'] or '',
        '租金总额': f2(d['租金总额']),
        '超时费': f2(d['超时费']),
        '损坏赔偿': f2(d['损坏赔偿']),
    })

# multiset 配对 (订单号, 商品名称)
csv_by_key = defaultdict(list)
for r in csv_rows: csv_by_key[(r['订单号'], r['商品名称'])].append(r)
xl_by_key = defaultdict(list)
for r in xl_rows: xl_by_key[(r['订单号'], r['租赁商品名称'])].append(r)

only_csv, only_xl = [], []
for key in set(csv_by_key) | set(xl_by_key):
    cl = list(csv_by_key.get(key, []))
    xl = list(xl_by_key.get(key, []))
    n = min(len(cl), len(xl))
    only_csv.extend(cl[n:])
    only_xl.extend(xl[n:])

only_csv.sort(key=lambda r: (r['订单号'], r['商品名称']))
only_xl.sort(key=lambda r: (r['订单号'], r['租赁商品名称']))

# 写
wb = Workbook()


def write_sheet(ws, title, color, headers, rows):
    ws.title = title
    ws.append(headers)
    fill = PatternFill('solid', fgColor=color)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    for row in rows:
        ws.append([row.get(h, '') for h in headers])
    ws.freeze_panes = 'A2'
    for c in range(1, len(headers) + 1):
        col = get_column_letter(c)
        max_len = sum(2 if ord(ch) > 127 else 1 for ch in str(headers[c - 1]))
        for r in range(2, min(len(rows), 200) + 2):
            v = ws.cell(row=r, column=c).value
            if v is None: continue
            s = str(v)
            wlen = sum(2 if ord(ch) > 127 else 1 for ch in s)
            if wlen > max_len: max_len = wlen
        ws.column_dimensions[col].width = min(max_len + 2, 36)


write_sheet(wb.active, '仅CSV有', 'C0392B',
            ['订单号', '订单类型', '商品名称', '订单日期', '订单时间',
             '起租日期', '退租日期', '结算日期', '租金', '_来源CSV'],
            only_csv)
write_sheet(wb.create_sheet(), '仅Excel有', '8E44AD',
            ['订单号', '租赁商品名称', '起租日期', '起租时间',
             '退租日期', '退租时间', '租金总额', '超时费', '损坏赔偿'],
            only_xl)
wb.save(OUT)
print(f'生成: {OUT}')
print(f'  仅CSV有: {len(only_csv)} 行  租金合计 {sum(r["租金"] for r in only_csv):,.2f}')
print(f'  仅Excel有: {len(only_xl)} 行  租金合计 {sum(r["租金总额"] for r in only_xl):,.2f}')
print(f'  文件大小: {os.path.getsize(OUT)/1024:.1f} KB')
