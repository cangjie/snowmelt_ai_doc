"""万龙体验中心 2025-10-15 ~ 2026-04-15 租赁订单 汇总+明细 导出"""
import pyodbc, sys, os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding='utf-8')

CONN = ('DRIVER={ODBC Driver 18 for SQL Server};SERVER=tcp:100.28.143.19,1433;'
        'DATABASE=snowmeet_new;UID=claude;PWD=abcd123!@#;'
        'Encrypt=yes;TrustServerCertificate=yes;Connection Timeout=30;')
SHOP = '万龙体验中心'
START = '2025-10-15'
END_EXCL = '2026-04-16'
OUT = r'D:\snowmeet\snowmeet_ai_doc\wanlong_rent_orders_2025-10-15_2026-04-15.xlsx'

# 退款条件：state=1 OR refund_id 非空（与 RentOrder.cs:519 旧代码一致）
REFUND_COND = "(pr.state = 1 OR (pr.refund_id IS NOT NULL AND pr.refund_id <> ''))"

SUMMARY_SQL = f"""
SELECT
    o.code                                              AS 订单号,
    CAST(o.biz_date AS DATE)                            AS 业务日期,
    CONVERT(VARCHAR(8), o.biz_date, 108)                AS 业务时间,
    CAST(refund_agg.last_refund_date AS DATE)           AS 结算日期,
    CONVERT(VARCHAR(8), refund_agg.last_refund_date, 108) AS 结算时间,
    ISNULL(pay_agg.paid_amount, 0)                      AS 支付总金额,
    ISNULL(refund_agg.refund_amount, 0)                 AS 退款总金额,
    ISNULL(pay_agg.paid_amount, 0) - ISNULL(refund_agg.refund_amount, 0) AS 订单结余,
    s.name                                              AS 店员姓名
FROM [order] o
LEFT JOIN staff s ON s.id = o.staff_id
LEFT JOIN (
    SELECT order_id, SUM(amount) AS paid_amount
    FROM order_payment
    WHERE status = N'支付成功' AND valid = 1
    GROUP BY order_id
) pay_agg ON pay_agg.order_id = o.id
LEFT JOIN (
    SELECT pr.order_id,
           SUM(pr.amount)            AS refund_amount,
           MAX(pr.create_date)       AS last_refund_date
    FROM payment_refund pr
    WHERE {REFUND_COND}
    GROUP BY pr.order_id
) refund_agg ON refund_agg.order_id = o.id
WHERE o.shop = ?
  AND o.[type] = N'租赁'
  AND o.create_date >= ?
  AND o.create_date < ?
  AND o.valid = 1
ORDER BY o.create_date ASC, o.id ASC
"""

DETAIL_SQL = f"""
SELECT
    o.code                                                                    AS 订单号,
    r.name                                                                    AS 租赁商品名称,
    CAST(r.start_date AS DATE)                                                AS 起租日期,
    CONVERT(VARCHAR(8), r.start_date, 108)                                    AS 起租时间,
    CAST(r.end_date AS DATE)                                                  AS 退租日期,
    CONVERT(VARCHAR(8), r.end_date, 108)                                      AS 退租时间,
    ISNULL(SUM(CASE WHEN rd.charge_type = N'租金'   AND rd.valid = 1 THEN rd.amount END), 0) AS 租金总额,
    ISNULL(SUM(CASE WHEN rd.charge_type = N'超时费' AND rd.valid = 1 THEN rd.amount END), 0) AS 超时费,
    ISNULL(SUM(CASE WHEN rd.charge_type = N'赔偿金' AND rd.valid = 1 THEN rd.amount END), 0) AS 损坏赔偿
FROM [order] o
JOIN rental r ON r.order_id = o.id AND r.valid = 1
LEFT JOIN rental_detail rd ON rd.rental_id = r.id
WHERE o.shop = ?
  AND o.[type] = N'租赁'
  AND o.create_date >= ?
  AND o.create_date < ?
  AND o.valid = 1
GROUP BY o.code, r.name, r.start_date, r.end_date, o.create_date, o.id, r.id
ORDER BY o.create_date ASC, o.id ASC, r.id ASC
"""

# 每笔支付一行；mch_id 来自 wepay_key（仅微信支付填，其他为空）
PAYMENT_SQL = f"""
SELECT
    o.code                                                  AS 订单号,
    op.pay_method                                           AS 支付方式,
    CASE WHEN op.pay_method = N'微信支付' THEN wk.mch_id ELSE NULL END AS mch_id,
    op.amount                                               AS 支付金额,
    ISNULL(rp.refund_amount, 0)                             AS 退款金额,
    op.amount - ISNULL(rp.refund_amount, 0)                 AS 结余金额
FROM [order] o
JOIN order_payment op ON op.order_id = o.id
LEFT JOIN wepay_key wk ON wk.id = op.mch_id
LEFT JOIN (
    SELECT pr.payment_id, SUM(pr.amount) AS refund_amount
    FROM payment_refund pr
    WHERE {REFUND_COND}
    GROUP BY pr.payment_id
) rp ON rp.payment_id = op.id
WHERE o.shop = ?
  AND o.[type] = N'租赁'
  AND o.create_date >= ?
  AND o.create_date < ?
  AND o.valid = 1
  AND op.status = N'支付成功'
  AND op.valid = 1
ORDER BY o.create_date ASC, o.id ASC, op.id ASC
"""


def write_sheet(ws, title, header_color, headers, rows):
    ws.title = title
    ws.append(headers)
    fill = PatternFill('solid', fgColor=header_color)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    for row in rows:
        ws.append(list(row))
    ws.freeze_panes = 'A2'
    for c in range(1, len(headers) + 1):
        col = get_column_letter(c)
        max_len = sum(2 if ord(ch) > 127 else 1 for ch in str(headers[c - 1]))
        for r in range(2, min(len(rows), 200) + 2):
            v = ws.cell(row=r, column=c).value
            if v is None: continue
            s = v.strftime('%Y-%m-%d %H:%M') if isinstance(v, datetime) else str(v)
            wlen = sum(2 if ord(ch) > 127 else 1 for ch in s)
            if wlen > max_len: max_len = wlen
        ws.column_dimensions[col].width = min(max_len + 2, 36)


def main():
    print('连数据库 ...')
    cn = pyodbc.connect(CONN)
    cur = cn.cursor()

    print(f'查汇总（万龙 / {START} ~ {END_EXCL}） ...')
    cur.execute(SUMMARY_SQL, SHOP, START, END_EXCL)
    summary_cols = [c[0] for c in cur.description]
    summary_rows = cur.fetchall()
    print(f'  汇总: {len(summary_rows)} 行')

    print('查明细 ...')
    cur.execute(DETAIL_SQL, SHOP, START, END_EXCL)
    detail_cols = [c[0] for c in cur.description]
    detail_rows = cur.fetchall()
    print(f'  明细: {len(detail_rows)} 行')

    print('查支付明细 ...')
    cur.execute(PAYMENT_SQL, SHOP, START, END_EXCL)
    pay_cols = [c[0] for c in cur.description]
    pay_rows = cur.fetchall()
    print(f'  支付明细: {len(pay_rows)} 行')

    cn.close()

    print(f'写 Excel: {OUT}')
    wb = Workbook()
    write_sheet(wb.active,        '订单汇总', '1F4E78', summary_cols, summary_rows)
    write_sheet(wb.create_sheet(), '订单明细', '2E7D32', detail_cols, detail_rows)
    write_sheet(wb.create_sheet(), '支付明细', 'B7791F', pay_cols,    pay_rows)
    wb.save(OUT)
    print(f'完成。文件大小: {os.path.getsize(OUT)/1024:.1f} KB')


if __name__ == '__main__':
    main()
