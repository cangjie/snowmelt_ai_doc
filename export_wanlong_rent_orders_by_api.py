"""万龙体验中心 2025-10-15 ~ 2026-04-15 租赁订单 — 通过本地 API 拉取（4 列汇总）

依赖本地 SnowmeetApi 服务（http://localhost:5099），不直连数据库。
3 个 2 个月非重叠窗口覆盖整个区间。
「总计租金」复刻小程序 pages/admin/rent/new_rent_list.js:236-251 的 displayedRental。
"""
import os
import sys
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding='utf-8')

BASE = "https://mini.snowmeet.top"
SESSION_KEY = "XqQs+sBFpWgIIIXPor7XmA=="
SHOP = "万龙体验中心"
TYPE = "租赁"
WINDOWS = [
    ("2025-10-15", "2025-12-14"),
    ("2025-12-15", "2026-02-14"),
    ("2026-02-15", "2026-04-15"),
]
OUT = "/Users/cangjie/source/snowmeet/snowmeet_ai/snowmeet_ai_doc/wanlong_rent_orders_api_2025-10-15_2026-04-15.xlsx"
HEADERS = ["订单号", "顾客称呼", "手机号", "总计租金"]


def compute_displayed_rental(o):
    """复刻 snowmeet_wechat_mini/pages/admin/rent/new_rent_list.js:236-251"""
    rentals = o.get("rentals") or []
    rent_props = o.get("rentProperties") or {}
    if rentals and rent_props.get("rentStatus") == "了结关闭":
        return sum(
            (r.get("totalRentalAmount") or 0) - (r.get("totalDiscountAmount") or 0)
            for r in rentals
        )
    if not rentals and (o.get("refundAmount") or 0) > 0:
        return (o.get("paidAmount") or 0) - (o.get("refundAmount") or 0)
    return 0


def fetch_window(start, end):
    params = {
        "sessionKey": SESSION_KEY,
        "type": TYPE,
        "shop": SHOP,
        "startDate": start,
        "endDate": end
    }
    r = requests.get(f"{BASE}/api/Order/GetOrdersByStaff", params=params, timeout=600)
    r.raise_for_status()
    body = r.json()
    if body.get("code") != 0:
        raise RuntimeError(f"接口报错 code={body.get('code')} message={body.get('message')}")
    return body.get("data") or []


def write_sheet(ws, rows):
    ws.title = "订单"
    ws.append(HEADERS)
    fill = PatternFill("solid", fgColor="1F4E78")
    for c in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in rows:
        ws.append(list(row))
    ws.freeze_panes = "A2"
    for c in range(1, len(HEADERS) + 1):
        col = get_column_letter(c)
        max_len = sum(2 if ord(ch) > 127 else 1 for ch in str(HEADERS[c - 1]))
        for r in range(2, min(len(rows), 200) + 2):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            s = str(v)
            wlen = sum(2 if ord(ch) > 127 else 1 for ch in s)
            if wlen > max_len:
                max_len = wlen
        ws.column_dimensions[col].width = min(max_len + 2, 36)


def main():
    rows = []
    for start, end in WINDOWS:
        print(f"拉窗口 {start} ~ {end} ...")
        orders = fetch_window(start, end)
        print(f"  返回 {len(orders)} 单")
        for o in orders:
            rows.append([
                o.get("code") or o.get("id"),
                o.get("customerCalledName") or "",
                o.get("customerCell") or "",
                compute_displayed_rental(o),
            ])
    rows.sort(key=lambda r: str(r[0]))
    print(f"合计 {len(rows)} 单")

    wb = Workbook()
    write_sheet(wb.active, rows)
    wb.save(OUT)
    size_kb = os.path.getsize(OUT) / 1024
    print(f"写入 {OUT}\n文件大小: {size_kb:.1f} KB")


if __name__ == "__main__":
    main()
