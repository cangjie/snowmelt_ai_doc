#!/usr/bin/env python3
"""一次性给财年版 xlsx 追加「支付明细」sheet。

订单号集合：取自目标 xlsx「年度租赁」sheet 的「订单号」列（不重做 DB 端 dedup）。
每笔成功支付一行；同一支付下的多笔退款用动态列内嵌（与主 sheet 风格一致）。

固定 10 列：
    订单号 / 支付订单号 / 支付方式 / 支付账户（微信=真实 mch_id / 支付宝='') / 顾客ID (微信 openid / 支付宝 ali_buyer_id)
    / 支付日期 / 支付时间 / 支付金额 / 退款金额 / 支付结余
动态 maxRefund × 4 列：
    退款k日期 / 退款k时间 / 退款k金额 / 退款k方式（=原支付方式）
动态 maxShare × 3 列：
    分账k金额 / 分账k成功（是/否/作废/''） / 分账k对象（order_share_relation.name）
    - 是：success=1
    - 否：success=0（请求被支付宝/微信驳回）
    - 作废：valid=0（请求生成后被软删，submit_time 一般为 null，未真实发出）
    - 空：success=null + valid=1（待处理）

幂等：sheet 已存在则删除重建。

用法：
    python3 add_payment_detail_sheet_to_fy_xlsx.py
    python3 add_payment_detail_sheet_to_fy_xlsx.py --xlsx other.xlsx --conn "..."

环境（macOS）：
    export ODBCSYSINI=/opt/homebrew/etc
"""
import argparse
import os
import sys
from collections import defaultdict
from datetime import datetime

os.environ.setdefault("ODBCSYSINI", "/opt/homebrew/etc")

import pyodbc  # noqa: E402
import openpyxl  # noqa: E402
from openpyxl.styles import Font, PatternFill, Alignment  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

sys.stdout.reconfigure(encoding="utf-8")

# 复用对账版 skill 的 REFUND_COND / DEFAULT_CONN / write_sheet
BASE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, os.path.join(BASE, "skills", "export_rent_order"))
from export_rent_orders import DEFAULT_CONN, REFUND_COND, write_sheet  # noqa: E402

DEFAULT_XLSX = os.path.join(
    BASE, "wanlong_rent_orders_fy_2025-05-01_2026-04-30.xlsx"
)
MAIN_SHEET = "年度租赁"
NEW_SHEET = "支付明细"
TX_SHEET = "支付流水"
KEY_HEADER = "订单号"

IN_BATCH = 2000  # SQL Server IN 上限 ~2100，留点 buffer


def read_order_codes(xlsx_path):
    """从主 sheet「年度租赁」读「订单号」列，返回 set（保留原始顺序作 sample 输出）。"""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if MAIN_SHEET not in wb.sheetnames:
        raise SystemExit(f"sheet '{MAIN_SHEET}' 不存在，实际: {wb.sheetnames}")
    ws = wb[MAIN_SHEET]
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    if KEY_HEADER not in header_row:
        raise SystemExit(
            f"表头缺「{KEY_HEADER}」: {header_row[:10]}..."
        )
    code_col_idx = header_row.index(KEY_HEADER)
    codes = []
    seen = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        c = row[code_col_idx]
        if c is None:
            continue
        s = str(c).strip()
        if not s or s in seen:
            continue
        seen.add(s)
        codes.append(s)
    wb.close()
    return codes


def fetch_payments(conn, order_codes):
    """按 code 分批 IN 查 order_payment，返回 list[dict]。"""
    pay_rows = []
    cur = conn.cursor()
    for i in range(0, len(order_codes), IN_BATCH):
        batch = order_codes[i : i + IN_BATCH]
        placeholders = ",".join(["?"] * len(batch))
        sql = f"""
            SELECT
                o.code                                                AS order_code,
                op.id                                                 AS payment_id,
                op.pay_method                                         AS pay_method,
                wk.mch_id                                             AS real_mch_id,
                COALESCE(NULLIF(op.open_id, ''), op.ali_buyer_id)     AS customer_id,
                op.paid_date                                          AS paid_date,
                op.create_date                                        AS create_date,
                op.amount                                             AS amount,
                op.out_trade_no                                       AS out_trade_no
            FROM [order] o
            JOIN order_payment op ON op.order_id = o.id
            LEFT JOIN wepay_key wk ON wk.id = op.mch_id
            WHERE op.status = N'支付成功'
              AND op.valid = 1
              AND o.code IN ({placeholders})
            ORDER BY o.code, COALESCE(op.paid_date, op.create_date), op.id
        """
        cur.execute(sql, *batch)
        for r in cur.fetchall():
            # 支付账户：微信支付=真实商户号（wepay_key.mch_id）；支付宝=''；其他保守填 ''
            if r.pay_method == "微信支付":
                pay_account = str(r.real_mch_id) if r.real_mch_id is not None else ""
            else:
                pay_account = ""
            pay_rows.append({
                "order_code": r.order_code,
                "payment_id": r.payment_id,
                "pay_method": r.pay_method,
                "pay_account": pay_account,
                "customer_id": r.customer_id or "",
                "paid_date": r.paid_date or r.create_date,  # paid_date NULL 时用 create_date 兜底
                "amount": float(r.amount or 0),
                "out_trade_no": r.out_trade_no or "",
            })
    return pay_rows


def fetch_shares(conn, payment_ids):
    """按 payment_id 分批 IN 查 payment_share + JOIN order_share + order_share_relation。
    返回 dict[payment_id → list[share_row]]。不过滤 valid（valid=0 用「作废」标签显示）。"""
    if not payment_ids:
        return {}
    share_by_pid = defaultdict(list)
    cur = conn.cursor()
    pid_list = list(payment_ids)
    for i in range(0, len(pid_list), IN_BATCH):
        batch = pid_list[i : i + IN_BATCH]
        placeholders = ",".join(["?"] * len(batch))
        sql = f"""
            SELECT
                ps.payment_id   AS payment_id,
                ps.amount       AS amount,
                ps.success      AS success,
                ps.valid        AS valid,
                ps.out_trade_no AS out_trade_no,
                ps.create_date  AS create_date,
                ps.response_time AS response_time,
                osr.name        AS payee_name
            FROM payment_share ps
            LEFT JOIN order_share os ON os.id = ps.share_id
            LEFT JOIN order_share_relation osr ON osr.id = os.relation_id
            WHERE ps.payment_id IN ({placeholders})
            ORDER BY ps.payment_id, ps.id
        """
        cur.execute(sql, *batch)
        for r in cur.fetchall():
            share_by_pid[r.payment_id].append({
                "amount": float(r.amount or 0),
                "success": r.success,  # bool? — True/False/None
                "valid": bool(r.valid),
                "out_trade_no": r.out_trade_no or "",
                "create_date": r.create_date,
                "response_time": r.response_time,
                "payee_name": r.payee_name or "",
            })
    return share_by_pid


def fetch_refunds(conn, payment_ids):
    """按 payment_id 分批 IN 查 payment_refund，返回 dict[payment_id → list[refund_row]]。"""
    if not payment_ids:
        return {}
    refund_by_pid = defaultdict(list)
    cur = conn.cursor()
    pid_list = list(payment_ids)
    for i in range(0, len(pid_list), IN_BATCH):
        batch = pid_list[i : i + IN_BATCH]
        placeholders = ",".join(["?"] * len(batch))
        sql = f"""
            SELECT
                pr.payment_id    AS payment_id,
                pr.amount        AS amount,
                pr.create_date   AS create_date,
                pr.out_refund_no AS out_refund_no
            FROM payment_refund pr
            WHERE pr.payment_id IN ({placeholders})
              AND {REFUND_COND}
            ORDER BY pr.payment_id, pr.create_date, pr.id
        """
        cur.execute(sql, *batch)
        for r in cur.fetchall():
            refund_by_pid[r.payment_id].append({
                "amount": float(r.amount or 0),
                "create_date": r.create_date,
                "out_refund_no": r.out_refund_no or "",
            })
    return refund_by_pid


def split_date(dt):
    if dt is None:
        return None
    if isinstance(dt, datetime):
        return dt.strftime("%Y-%m-%d")
    return str(dt)[:10]


def split_time(dt):
    if dt is None:
        return None
    if isinstance(dt, datetime):
        return dt.strftime("%H:%M:%S")
    return None


def _success_label(success, valid):
    """payment_share.(success, valid) → 显示标签
    - valid=0 → 作废（请求被软删，未真实发出）
    - valid=1 且 success=1 → 是
    - valid=1 且 success=0 → 否（接口驳回）
    - valid=1 且 success=None → ''（待处理）
    """
    if not valid:
        return "作废"
    if success is True:
        return "是"
    if success is False:
        return "否"
    return ""


def build_headers_and_rows(pay_rows, refund_by_pid, share_by_pid):
    maxRefund = max(
        (len(refund_by_pid.get(p["payment_id"], [])) for p in pay_rows),
        default=0,
    )
    maxShare = max(
        (len(share_by_pid.get(p["payment_id"], [])) for p in pay_rows),
        default=0,
    )

    headers = [
        "订单号", "支付订单号", "支付方式", "支付账户", "顾客ID",
        "支付日期", "支付时间", "支付金额", "退款金额", "支付结余",
    ]
    for k in range(1, maxRefund + 1):
        headers += [f"退款{k}日期", f"退款{k}时间", f"退款{k}金额", f"退款{k}方式"]
    for k in range(1, maxShare + 1):
        headers += [f"分账{k}金额", f"分账{k}成功", f"分账{k}对象"]

    rows = []
    for p in pay_rows:
        refunds = refund_by_pid.get(p["payment_id"], [])
        refund_sum = round(sum(r["amount"] for r in refunds), 2)
        balance = round(p["amount"] - refund_sum, 2)  # 支付结余 = 支付金额 − 退款金额
        shares = share_by_pid.get(p["payment_id"], [])
        row = [
            p["order_code"], p["payment_id"], p["pay_method"],
            p["pay_account"], p["customer_id"],
            split_date(p["paid_date"]), split_time(p["paid_date"]),
            round(p["amount"], 2), refund_sum, balance,
        ]
        for k in range(maxRefund):
            if k < len(refunds):
                r = refunds[k]
                row += [
                    split_date(r["create_date"]),
                    split_time(r["create_date"]),
                    round(r["amount"], 2),
                    p["pay_method"],  # 退款方式 = 原支付通道（payment_refund 无 pay_method 列）
                ]
            else:
                row += [None, None, None, None]
        for k in range(maxShare):
            if k < len(shares):
                s = shares[k]
                row += [round(s["amount"], 2), _success_label(s["success"], s["valid"]), s["payee_name"]]
            else:
                row += [None, None, None]
        rows.append(row)

    return headers, rows, maxRefund, maxShare


def build_transaction_rows(pay_rows, refund_by_pid, share_by_pid):
    """支付流水行（按 日期+时间 升序）。
    - 支付：order_payment 全部 pay_rows（已是 status=支付成功 valid=1），交易金额取正
    - 退款：payment_refund 命中 REFUND_COND（state=1 OR refund_id 非空），交易金额取负
    - 分账：payment_share success=1 AND valid=1，交易金额取负（站在商家本主体角度算流出）
    支付方式/支付账户：退款/分账继承自所属 payment（payment_refund 无 pay_method 列；分账走原通道）
    """
    headers = ["订单号", "支付方式", "支付账户", "商户订单号", "类型", "交易金额", "日期", "时间"]
    pid_ctx = {p["payment_id"]: p for p in pay_rows}
    rows = []

    # 支付（正）
    for p in pay_rows:
        rows.append({
            "订单号": p["order_code"],
            "支付方式": p["pay_method"],
            "支付账户": p["pay_account"],
            "商户订单号": p["out_trade_no"],
            "类型": "支付",
            "金额": round(p["amount"], 2),
            "_dt": p["paid_date"],
        })

    # 退款（负）
    for pid, rfs in refund_by_pid.items():
        parent = pid_ctx.get(pid)
        for r in rfs:
            rows.append({
                "订单号": parent["order_code"] if parent else "",
                "支付方式": parent["pay_method"] if parent else "",
                "支付账户": parent["pay_account"] if parent else "",
                "商户订单号": r["out_refund_no"],
                "类型": "退款",
                "金额": -round(r["amount"], 2),
                "_dt": r["create_date"],
            })

    # 分账（负 — 站在商家本主体角度是流出）
    for pid, shares in share_by_pid.items():
        parent = pid_ctx.get(pid)
        for s in shares:
            if s["success"] is not True or not s["valid"]:
                continue
            rows.append({
                "订单号": parent["order_code"] if parent else "",
                "支付方式": parent["pay_method"] if parent else "",
                "支付账户": parent["pay_account"] if parent else "",
                "商户订单号": s["out_trade_no"],
                "类型": "分账",
                "金额": -round(s["amount"], 2),
                "_dt": s["response_time"] or s["create_date"],
            })

    # 按 日期+时间 升序（None 落最前）
    rows.sort(key=lambda r: (r["_dt"] is None, r["_dt"]))

    final_rows = []
    for r in rows:
        final_rows.append([
            r["订单号"], r["支付方式"], r["支付账户"], r["商户订单号"], r["类型"],
            r["金额"],
            split_date(r["_dt"]), split_time(r["_dt"]),
        ])
    return headers, final_rows


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", default=DEFAULT_XLSX, help="目标 xlsx 路径")
    parser.add_argument("--conn", default=DEFAULT_CONN, help="ODBC 连接串")
    args = parser.parse_args()

    if not os.path.exists(args.xlsx):
        raise SystemExit(f"xlsx 不存在: {args.xlsx}")

    print(f"读取主 sheet「{MAIN_SHEET}」订单号集合 ...")
    order_codes = read_order_codes(args.xlsx)
    print(f"  共 {len(order_codes)} 个订单号；sample: {order_codes[:3]}")

    print(f"连接 DB ...")
    conn = pyodbc.connect(args.conn)

    print(f"查询 order_payment ...")
    pay_rows = fetch_payments(conn, order_codes)
    print(f"  共 {len(pay_rows)} 笔成功支付")

    payment_ids = {p["payment_id"] for p in pay_rows}
    print(f"查询 payment_refund ...")
    refund_by_pid = fetch_refunds(conn, payment_ids)
    n_refund_total = sum(len(v) for v in refund_by_pid.values())
    print(f"  共 {n_refund_total} 笔有效退款（命中 REFUND_COND），分布在 {len(refund_by_pid)} 笔支付下")

    print(f"查询 payment_share ...")
    share_by_pid = fetch_shares(conn, payment_ids)
    n_share_total = sum(len(v) for v in share_by_pid.values())
    n_voided = sum(1 for v in share_by_pid.values() for s in v if not s["valid"])
    print(f"  共 {n_share_total} 笔分账记录（含 {n_voided} 笔 valid=0 作废），分布在 {len(share_by_pid)} 笔支付下")
    conn.close()

    headers, rows, maxRefund, maxShare = build_headers_and_rows(pay_rows, refund_by_pid, share_by_pid)
    print(f"\n输出列：{len(headers)}（固定 10 + 退款 {maxRefund}×4 + 分账 {maxShare}×3）")
    print(f"输出行：{len(rows)}")

    # 打开目标 xlsx，幂等地删旧 sheet 重建
    wb = openpyxl.load_workbook(args.xlsx)
    if NEW_SHEET in wb.sheetnames:
        print(f"删除已存在的「{NEW_SHEET}」sheet")
        del wb[NEW_SHEET]
    ws = wb.create_sheet(NEW_SHEET)

    # 复用 export_rent_orders.write_sheet（粗白字 + 1F4E78 蓝底 header + freeze A2 + 自适应列宽）
    write_sheet(ws, NEW_SHEET, "1F4E78", headers, rows)

    # 金额列锁定 0.00 显示格式（避开科学计数法）
    money_col_idxs = [8, 9, 10]  # 支付金额、退款金额、支付结余（1-based）
    refund_block_end = 10 + maxRefund * 4
    for k in range(maxRefund):
        money_col_idxs.append(10 + k * 4 + 3)  # 退款k金额
    for k in range(maxShare):
        money_col_idxs.append(refund_block_end + k * 3 + 1)  # 分账k金额
    for col_idx in money_col_idxs:
        for row_idx in range(2, len(rows) + 2):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                cell.number_format = "0.00"

    # === 「支付流水」sheet ===
    tx_headers, tx_rows = build_transaction_rows(pay_rows, refund_by_pid, share_by_pid)
    cnt_pay = sum(1 for r in tx_rows if r[4] == '支付')
    cnt_refund = sum(1 for r in tx_rows if r[4] == '退款')
    cnt_share = sum(1 for r in tx_rows if r[4] == '分账')
    sum_signed = sum(r[5] for r in tx_rows if r[5] is not None)
    print(f"\n支付流水 sheet：{len(tx_rows)} 行（支付 {cnt_pay} / 退款 {cnt_refund} / 分账 {cnt_share}）")
    print(f"  交易金额合计（含符号）: ¥{sum_signed:.2f}")

    # 清理可能存在的旧名 sheet
    for old_name in ("成功交易", TX_SHEET):
        if old_name in wb.sheetnames:
            print(f"删除已存在的「{old_name}」sheet")
            del wb[old_name]
    tx_ws = wb.create_sheet(TX_SHEET)
    write_sheet(tx_ws, TX_SHEET, "1F4E78", tx_headers, tx_rows)

    # 交易金额列锁定 0.00 显示格式
    amount_col = tx_headers.index("交易金额") + 1
    for row_idx in range(2, len(tx_rows) + 2):
        cell = tx_ws.cell(row=row_idx, column=amount_col)
        if cell.value is not None:
            cell.number_format = "0.00"

    wb.save(args.xlsx)

    size_kb = os.path.getsize(args.xlsx) / 1024
    print(f"\n写入 {args.xlsx}")
    print(f"文件大小: {size_kb:.1f} KB")


if __name__ == "__main__":
    main()
