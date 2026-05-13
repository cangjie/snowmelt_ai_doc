"""把 csv_excel_diff.xlsx 中「仅Excel有」按接口 fail 原因拆成多个 sheet"""
import sys, pyodbc, os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding='utf-8')

FILE = r'D:\snowmeet\snowmeet_ai_doc\csv_excel_diff.xlsx'
CONN = ('DRIVER={ODBC Driver 18 for SQL Server};SERVER=tcp:100.28.143.19,1433;'
        'DATABASE=snowmeet_new;UID=claude;PWD=abcd123!@#;'
        'Encrypt=yes;TrustServerCertificate=yes;Connection Timeout=30;')

# 接口规则 (RentController.cs:5544 GetConfirmedRentOrder)
# 返回需同时：paidAmount > 0 AND closed=1 AND close_date != null AND !hide
#            AND 不含非微信非支付宝的支付
SHEETS = [
    ('paid为0',          'C0392B'),
    ('closed为0_未关闭',   'E67E22'),
    ('close_date为空',   '7F8C8D'),
    ('hide为1_隐藏',      '34495E'),
    ('含非微信非支付宝',     '16A085'),
    ('应通过但CSV没有',     '2980B9'),
]


def classify(closed, close_date, hide, paid, nonwx):
    """按接口规则顺序判断 fail 原因；通过则归到「应通过但CSV没有」"""
    if not (paid and paid > 0):
        return 'paid为0'
    if closed != 1:
        return 'closed为0_未关闭'
    if close_date is None:
        return 'close_date为空'
    if hide:
        return 'hide为1_隐藏'
    if nonwx and nonwx > 0:
        return '含非微信非支付宝'
    return '应通过但CSV没有'


# 加载现有文件
wb = load_workbook(FILE)
ws_excel = wb['仅Excel有']
hdr = [c.value for c in ws_excel[1]]
rows = []
for r in ws_excel.iter_rows(min_row=2, values_only=True):
    rows.append(dict(zip(hdr, r)))
codes = list({r['订单号'] for r in rows if r['订单号']})
print(f'仅Excel有: {len(rows)} 行明细，{len(codes)} 个去重订单')

# 查这些订单的接口判定字段
cn = pyodbc.connect(CONN)
cur = cn.cursor()
in_clause = ','.join(['?'] * len(codes))
cur.execute(f"""
SELECT o.code, o.closed, o.close_date, o.hide,
  (SELECT ISNULL(SUM(amount),0) FROM order_payment WHERE order_id=o.id AND status=N'支付成功' AND valid=1) AS paid,
  (SELECT COUNT(*) FROM order_payment WHERE order_id=o.id AND status=N'支付成功' AND valid=1
    AND pay_method NOT IN (N'微信支付', N'支付宝')) AS nonwx
FROM [order] o WHERE o.code IN ({in_clause})
""", *codes)
order_meta = {}
for r in cur.fetchall():
    order_meta[r[0]] = {
        'closed': r[1], 'close_date': r[2], 'hide': r[3], 'paid': r[4], 'nonwx': r[5],
        'reason': classify(r[1], r[2], r[3], r[4], r[5]),
    }
cn.close()

# 移除可能已存在的旧分类 sheet（重跑幂等）
for name, _ in SHEETS:
    if name in wb.sheetnames:
        del wb[name]

# 分类装行
bucket = {name: [] for name, _ in SHEETS}
for r in rows:
    code = r['订单号']
    meta = order_meta.get(code)
    reason = meta['reason'] if meta else 'close_date为空'  # fallback
    # 给行附加额外说明列（订单状态快照）
    enriched = dict(r)
    if meta:
        enriched['_订单已付金额'] = round(meta['paid'] or 0, 2)
        enriched['_订单closed'] = meta['closed']
        enriched['_订单close_date'] = meta['close_date'].strftime('%Y-%m-%d %H:%M') if meta['close_date'] else ''
        enriched['_订单hide'] = '是' if meta['hide'] else ''
        enriched['_非微信非支付宝笔数'] = meta['nonwx']
    bucket[reason].append(enriched)


def write_sheet(ws, title, color, headers, rows):
    ws.title = title
    ws.append(headers)
    fill = PatternFill('solid', fgColor=color)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    for row in rows:
        ws.append([row.get(h, '') for h in headers])
    ws.freeze_panes = 'A2'
    for c in range(1, len(headers) + 1):
        col = get_column_letter(c)
        max_len = sum(2 if ord(ch) > 127 else 1 for ch in str(headers[c - 1]))
        for r in range(2, min(len(rows), 200) + 2):
            v = ws.cell(row=r, column=c).value
            if v is None: continue
            s = str(v)
            wlen = sum(2 if ord(ch) > 127 else 1 for ch in s)
            if wlen > max_len: max_len = wlen
        ws.column_dimensions[col].width = min(max_len + 2, 36)


sheet_headers = hdr + ['_订单已付金额', '_订单closed', '_订单close_date', '_订单hide', '_非微信非支付宝笔数']
for name, color in SHEETS:
    ws = wb.create_sheet(name)
    write_sheet(ws, name, color, sheet_headers, bucket[name])

wb.save(FILE)

print(f'\n已写入 {FILE}')
for name, _ in SHEETS:
    rs = bucket[name]
    uniq = len({r['订单号'] for r in rs})
    rent = sum(r.get('租金总额', 0) or 0 for r in rs)
    print(f'  {name:25s} 明细 {len(rs):4d} 行  订单 {uniq:4d} 个  租金合计 {rent:>14,.2f}')
print(f'  文件大小: {os.path.getsize(FILE)/1024:.1f} KB')
