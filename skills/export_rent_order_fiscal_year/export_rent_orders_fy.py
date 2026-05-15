#!/usr/bin/env python3
"""按店铺 + 财年口径导出年度租赁订单（单 sheet 宽表，财务/业务视角）。

与同目录 ../export_rent_order（3 sheet 对账版）互补：
- 本 skill 单 sheet，5 段拼接：固定前缀 + 动态支付区 + 动态退款区 + 固定中段 + 固定后缀
- 支付/退款明细列数 = 导出区间内单订单最大成功支付 / 有效退款笔数（数据驱动，每次跑可能不同）
- 财年/营非/运营日序号按 biz_date 派生；营业区间逐财年内置 SEASON dict（缺失财年报错提示补表）

用法：
    python export_rent_orders_fy.py --shop 万龙体验中心
    python export_rent_orders_fy.py --shop 渔阳 --start 2025-05-01 --end 2026-04-30 --out yuyang_fy.xlsx

环境（Windows）：pip install pyodbc openpyxl；需安装 "ODBC Driver 18 for SQL Server"。
日期范围按 order.biz_date（业务日期）过滤——年度/财年报表语义：导的是「业务发生在该
财年」的订单。注意与 ../export_rent_order（按 create_date 下单时间过滤）口径不同，
两者结果不可 1:1 交叉对账。财年/营非/运营日序号亦按 biz_date 派生。

数据源默认硬编码生产（复用 ../export_rent_order/DEFAULT_CONN），可用 --conn 覆盖。
"""
import argparse
import os
import sys
from collections import defaultdict
from datetime import datetime, timedelta, date

import pyodbc
from openpyxl import Workbook

# 复用 ../export_rent_order 的单点真理（SHOP_PREFIX / REFUND_COND / DEFAULT_CONN / write_sheet）
sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'export_rent_order'))
try:
    from export_rent_orders import SHOP_PREFIX, REFUND_COND, DEFAULT_CONN, write_sheet
except ImportError as e:
    print(f'无法 import ../export_rent_order/export_rent_orders.py：{e}\n'
          f'本 skill 依赖 sibling 目录 export_rent_order，请确认两个 skill 目录关系未被破坏。',
          file=sys.stderr)
    raise

sys.stdout.reconfigure(encoding='utf-8')

# ───────────────────────── 财年营业区间表（逐财年，缺失报错提示补） ─────────────────────────
# key = 财年标签 'YY-YY'（biz_date 落在 5/1~次年4/30）；value = (雪季营业起, 雪季营业止) inclusive
# 仅用于「营/非」「运营日序号」派生；导其它财年请在此补行。
SEASON = {
    '25-26': (date(2025, 10, 21), date(2026, 4, 9)),
}

SHEET_TITLE = '年度租赁'
HEADER_COLOR = '1F4E78'

# ───────────────────────── SQL ─────────────────────────
# 通用过滤：单店 + 租赁 + 业务日期区间 + code 非空（+ 可选 order.valid=1）。
# code 非空恒过滤：无 code 的多是未下单/废弃单（从未生成订单号），不是真实业务记录，
#   不进财务报表（即便 --include-invalid 也排除）。
# __VALID__ 是运行期占位（非 f-string 插值，故能穿过下方 f-string 保留到 main 替换）：
#   默认 → 'AND o.valid = 1'（与对账版可比）；--include-invalid → ''（不论 valid 全导）
# 注意：valid 放宽仅作用于 order 表；rental/order_payment/discount/order_share 等不变。
ORDER_FILTER = (
    "o.shop = ? AND o.[type] = N'租赁' "
    "AND o.biz_date >= ? AND o.biz_date < ? "
    "AND o.code IS NOT NULL AND LTRIM(RTRIM(o.code)) <> N'' __VALID__"
)

# maxPay/maxRefund 不再用预查询：去重后按实际保留订单的支付/退款明细在 Python 端取 max
# 主查询：一行一订单。段1(部分)+段4+段5 的订单级列，全部用聚合/标量子查询保订单粒度。
MAIN_SQL = f"""
SELECT
    o.id                                                   AS _oid,
    o.valid                                                AS _valid,
    o.code                                                 AS 订单号,
    o.biz_date                                             AS _biz_dt,
    o.create_date                                          AS _crt_dt,
    ROW_NUMBER() OVER (PARTITION BY o.shop, CAST(o.biz_date AS DATE)
                       ORDER BY o.create_date ASC, o.id ASC) AS 日序号,
    ISNULL(pay_agg.cnt, 0)                                  AS 支付次数,
    ISNULL(pay_agg.paid, 0)                                 AS 支付合计,
    ISNULL(ref_agg.cnt, 0)                                  AS 退款次数,
    ISNULL(ref_agg.refund, 0)                               AS 退款合计,
    ISNULL(pay_agg.paid, 0) - ISNULL(ref_agg.refund, 0)     AS 订单结余,
    ref_agg.last_refund_dt                                  AS _last_refund_dt,
    ot.overtime                                             AS 超时费合计,
    ot.damage                                               AS 赔偿合计,
    ISNULL((
        SELECT SUM(dd.amount) FROM (
            SELECT DISTINCT d.id, d.amount
            FROM discount d
            WHERE d.valid = 1 AND (
                  d.order_id = o.id
               OR (d.biz_type = N'租赁' AND d.biz_id IN (
                     SELECT r.id FROM rental r WHERE r.order_id = o.id AND r.valid = 1))
               OR (d.sub_biz_type IN (N'日租金', N'租赁项') AND d.sub_biz_id IN (
                     SELECT rd.id FROM rental_detail rd
                     JOIN rental r ON r.id = rd.rental_id
                     WHERE r.order_id = o.id AND r.valid = 1))
            )
        ) dd
    ), 0)                                                   AS 减免合计,
    CASE WHEN o.hide = 1 THEN N'是' ELSE N'' END            AS 隐藏订单,
    ISNULL(osh.should_share, 0)                             AS 应分账金额,
    ISNULL(psh.actual_share, 0)                             AS 实分账金额,
    ISNULL(osh.should_share, 0) - ISNULL(psh.actual_share, 0) AS 待分账金额,
    o.shop                                                  AS 门店,
    COALESCE(NULLIF(LTRIM(RTRIM(o.contact_name)), N''), m.real_name) AS 客户名称,
    COALESCE(NULLIF(LTRIM(RTRIM(o.contact_num)), N''), msa_cell.num) AS 电话,
    msa_uid.num                                             AS [union id],
    big_pay.pay_method                                      AS 收款方式,
    CASE WHEN big_pay.pay_method = N'微信支付' THEN wk.mch_id ELSE NULL END AS 支付账号,
    CAST(o.biz_date AS DATE)                                AS 业务日期,
    CONVERT(VARCHAR(8), o.biz_date, 108)                    AS 业务时间,
    s.name                                                  AS 店员姓名,
    CASE WHEN ISNULL(pay_agg.paid, 0) < 5
              OR (s.name IS NOT NULL AND s.name LIKE N'%苍%')
         THEN N'是' ELSE N'' END                            AS 测试,
    ISNULL(rc.rental_cnt, 0)                                AS _rental_cnt,
    ISNULL(rc.settled_cnt, 0)                               AS _settled_cnt,
    rc.min_start                                            AS _min_start,
    ISNULL(rc.any_end_null, 0)                              AS _any_end_null,
    ISNULL(rc.rent_summary, 0)                              AS _rent_summary,
    o.closed                                                AS _closed,
    CASE WHEN o.entertain = 1 OR EXISTS (
             SELECT 1 FROM rental r WHERE r.order_id = o.id
                                      AND r.valid = 1 AND r.entertain = 1)
         THEN 1 ELSE 0 END                                  AS _entertain
FROM [order] o
LEFT JOIN staff s  ON s.id = o.staff_id
LEFT JOIN member m ON m.id = o.member_id
LEFT JOIN (
    SELECT order_id, COUNT(*) AS cnt, SUM(amount) AS paid
    FROM order_payment WHERE status = N'支付成功' AND valid = 1
    GROUP BY order_id
) pay_agg ON pay_agg.order_id = o.id
LEFT JOIN (
    SELECT pr.order_id, COUNT(*) AS cnt, SUM(pr.amount) AS refund,
           MAX(pr.create_date) AS last_refund_dt
    FROM payment_refund pr WHERE {REFUND_COND}
    GROUP BY pr.order_id
) ref_agg ON ref_agg.order_id = o.id
LEFT JOIN (
    SELECT r.order_id,
           ISNULL(SUM(CASE WHEN rd.charge_type = N'超时费' AND rd.valid = 1 THEN rd.amount END), 0) AS overtime,
           ISNULL(SUM(CASE WHEN rd.charge_type IN (N'赔偿金', N'损坏赔偿') AND rd.valid = 1 THEN rd.amount END), 0) AS damage
    FROM rental r LEFT JOIN rental_detail rd ON rd.rental_id = r.id
    WHERE r.valid = 1
    GROUP BY r.order_id
) ot ON ot.order_id = o.id
LEFT JOIN (
    SELECT order_id, SUM(amount) AS should_share
    FROM order_share WHERE valid = 1 GROUP BY order_id
) osh ON osh.order_id = o.id
LEFT JOIN (
    SELECT os2.order_id, SUM(ps.amount) AS actual_share
    FROM payment_share ps JOIN order_share os2 ON os2.id = ps.share_id
    WHERE ps.valid = 1 AND ps.success = 1
    GROUP BY os2.order_id
) psh ON psh.order_id = o.id
LEFT JOIN (
    SELECT r.order_id,
           COUNT(*) AS rental_cnt,
           SUM(CASE WHEN r.settled = 1 THEN 1 ELSE 0 END) AS settled_cnt,
           MIN(r.start_date) AS min_start,
           MAX(CASE WHEN r.end_date IS NULL THEN 1 ELSE 0 END) AS any_end_null,
           ISNULL((SELECT SUM(rd2.amount) FROM rental_detail rd2
                   JOIN rental r2 ON r2.id = rd2.rental_id
                   WHERE r2.order_id = r.order_id AND r2.valid = 1
                     AND rd2.charge_type = N'租金' AND rd2.valid = 1), 0) AS rent_summary
    FROM rental r WHERE r.valid = 1
    GROUP BY r.order_id
) rc ON rc.order_id = o.id
OUTER APPLY (
    SELECT TOP 1 op.pay_method, op.mch_id
    FROM order_payment op
    WHERE op.order_id = o.id AND op.status = N'支付成功' AND op.valid = 1
    ORDER BY op.amount DESC, op.id ASC
) big_pay
LEFT JOIN wepay_key wk ON wk.id = big_pay.mch_id
OUTER APPLY (
    SELECT TOP 1 num FROM member_social_account
    WHERE member_id = o.member_id AND type = N'cell' AND valid = 1
      AND LTRIM(RTRIM(num)) <> N'' ORDER BY id ASC
) msa_cell
OUTER APPLY (
    SELECT TOP 1 num FROM member_social_account
    WHERE member_id = o.member_id AND type = N'wechat_unionid' AND valid = 1
      AND LTRIM(RTRIM(num)) <> N'' ORDER BY id ASC
) msa_uid
WHERE {ORDER_FILTER}
ORDER BY o.biz_date ASC, o.create_date ASC, o.id ASC
"""

# 支付明细（段2）：每笔成功有效支付，按支付时间升序
PAY_DETAIL_SQL = f"""
SELECT o.id AS oid,
       COALESCE(op.paid_date, op.create_date) AS pt,
       op.amount,
       op.pay_method,
       CASE WHEN op.pay_method = N'微信支付' THEN wk.mch_id ELSE NULL END AS acct
FROM [order] o
JOIN order_payment op ON op.order_id = o.id
LEFT JOIN wepay_key wk ON wk.id = op.mch_id
WHERE {ORDER_FILTER} AND op.status = N'支付成功' AND op.valid = 1
ORDER BY o.id ASC, COALESCE(op.paid_date, op.create_date) ASC, op.id ASC
"""

# 退款明细（段3）：每笔有效退款，退款方式取原支付通道 pay_method（payment_refund 无方式列）
REFUND_DETAIL_SQL = f"""
SELECT o.id AS oid,
       pr.create_date AS rt,
       pr.amount,
       op.pay_method AS refund_method
FROM [order] o
JOIN payment_refund pr ON pr.order_id = o.id
LEFT JOIN order_payment op ON op.id = pr.payment_id
WHERE {ORDER_FILTER} AND {REFUND_COND}
ORDER BY o.id ASC, pr.create_date ASC, pr.id ASC
"""

# 主查询里以 _ 前缀的辅助列（不进 Excel，仅供 Python 派生）
AUX_COLS = {'_oid', '_valid', '_biz_dt', '_crt_dt', '_last_refund_dt', '_rental_cnt',
            '_settled_cnt', '_min_start', '_any_end_null', '_rent_summary',
            '_closed', '_entertain'}

# 需要 round(2) + number_format='0.00' 的金额列名（固定列部分；动态列另判）
MONEY_FIXED = {'支付合计', '退款合计', '订单结余', '超时费合计', '赔偿合计',
               '减免合计', '应分账金额', '实分账金额', '待分账金额'}


def fiscal_year(d):
    """biz_date → 财年标签 'YY-YY'（5/1 起，次年 4/30 止）"""
    y = d.year
    fy = y if d.month >= 5 else y - 1
    return f'{fy % 100:02d}-{(fy + 1) % 100:02d}'


def split_dt(v):
    """datetime → (date, 'HH:MM:SS')；None → (None, None)"""
    if v is None:
        return None, None
    return v.date(), v.strftime('%H:%M:%S')


def derive_status(rental_cnt, settled_cnt, min_start, any_end_null,
                  paid, refund, rent_summary, closed):
    """rentProperties.rentStatus 的 SQL 化近似（见 SKILL.md「订单状态」说明）。
    复刻 Order.cs:1134-1172 的判定顺序（后置条件覆盖前置）。realStart/totalSummary
    等计算属性无法纯 SQL 复现，用 start_date / 租金合计 近似。"""
    if rental_cnt == 0:
        return ''  # 后端 rentProperties 为 null
    today = date.today()
    ms = min_start.date() if isinstance(min_start, datetime) else min_start
    if ms is None or ms > today:
        return '未开始'
    status = ''
    if any_end_null:
        status = '租赁中'
    if 0 < settled_cnt < rental_cnt:
        status = '部分归还'
    if settled_cnt == rental_cnt and rental_cnt > 0:
        status = '全部归还'
    if refund and refund > 0:
        if (paid - rent_summary) <= refund:
            status = '了结关闭' if closed == 1 else '全额退押金'
        else:
            status = '部分退押金'
    return status


def parse_args():
    p = argparse.ArgumentParser(
        description='按店铺+财年导出年度租赁订单（单 sheet 宽表）',
        formatter_class=argparse.RawDescriptionHelpFormatter, epilog=__doc__)
    p.add_argument('--shop', required=True, help='店铺名（DB order.shop，如「万龙体验中心」）')
    p.add_argument('--start', default='2025-05-01', help='业务起始日期 biz_date inclusive，默认 2025-05-01')
    p.add_argument('--end', default='2026-04-30', help='业务截止日期 biz_date inclusive，默认 2026-04-30')
    p.add_argument('--out', default=None, help='输出 xlsx 路径，默认 {prefix}_rent_orders_fy_{start}_{end}.xlsx')
    p.add_argument('--conn', default=DEFAULT_CONN, help='ODBC 连接串，默认连生产')
    p.add_argument('--include-invalid', action='store_true',
                   help='导出 order 不论 valid 是否=1（默认仅 valid=1）。'
                        '仅放宽 order 表；rental/支付/退款/分账等 valid 过滤不变。'
                        '开启后与对账版口径不一致，不可交叉对账')
    return p.parse_args()


def default_out_name(shop, start, end):
    prefix = SHOP_PREFIX.get(shop, shop)
    return f'{prefix}_rent_orders_fy_{start}_{end}.xlsx'


def main():
    args = parse_args()
    start = args.start
    end_excl = (datetime.strptime(args.end, '%Y-%m-%d') + timedelta(days=1)).strftime('%Y-%m-%d')
    out = os.path.abspath(args.out or default_out_name(args.shop, args.start, args.end))
    params = (args.shop, start, end_excl)

    # 运行期替换 order.valid 占位（仅作用于 order 表过滤）
    valid_clause = '' if args.include_invalid else 'AND o.valid = 1'
    main_sql = MAIN_SQL.replace('__VALID__', valid_clause)
    pay_detail_sql = PAY_DETAIL_SQL.replace('__VALID__', valid_clause)
    refund_detail_sql = REFUND_DETAIL_SQL.replace('__VALID__', valid_clause)
    print(f'order.valid 过滤: {"关闭（不论 valid 全导）" if args.include_invalid else "valid=1"}')

    print('连数据库 ...')
    cn = pyodbc.connect(args.conn)
    cur = cn.cursor()

    print(f'主查询（{args.shop} / 业务 {start} ~ {end_excl}） ...')
    cur.execute(main_sql, *params)
    main_cols = [c[0] for c in cur.description]
    main_rows = cur.fetchall()
    print(f'  订单: {len(main_rows)} 行')

    print('支付明细 ...')
    cur.execute(pay_detail_sql, *params)
    pay_by_oid = defaultdict(list)
    for oid, pt, amt, pm, acct in cur.fetchall():
        pay_by_oid[oid].append((pt, amt, pm, acct))
    print(f'  支付笔: {sum(len(v) for v in pay_by_oid.values())}')

    print('退款明细 ...')
    cur.execute(refund_detail_sql, *params)
    ref_by_oid = defaultdict(list)
    for oid, rt, amt, rm in cur.fetchall():
        ref_by_oid[oid].append((rt, amt, rm))
    print(f'  退款笔: {sum(len(v) for v in ref_by_oid.values())}')
    cn.close()

    idx = {name: i for i, name in enumerate(main_cols)}
    I_OID, I_VALID = idx['_oid'], idx['_valid']
    I_CODE, I_PAYCNT = idx['订单号'], idx['支付次数']

    # ── 同订单号去重（用户指定优先级）──
    # 1) 有成功支付记录的那条；2) 否则 valid=1 的那条；3) 否则 id 最大的那条
    groups = defaultdict(list)
    for r in main_rows:
        groups[str(r[I_CODE]).strip()].append(r)
    deduped, dropped, dup_codes = [], 0, 0
    for code, grp in groups.items():
        if len(grp) > 1:
            dup_codes += 1
            dropped += len(grp) - 1
            winner = max(grp, key=lambda r: (
                1 if (r[I_PAYCNT] or 0) > 0 else 0,   # 有成功支付优先
                1 if r[I_VALID] == 1 else 0,           # 再 valid=1 优先
                r[I_OID]))                             # 再 id 最大
            deduped.append(winner)
        else:
            deduped.append(grp[0])
    deduped.sort(key=lambda r: (r[idx['_biz_dt']], r[idx['_crt_dt']], r[I_OID]))
    print(f'  重复订单号 {dup_codes} 个，去重丢弃 {dropped} 行 → 保留 {len(deduped)} 行')
    main_rows = deduped

    # 动态支付/退款列数按去重后实际保留集计算（精确，不依赖预查询）
    kept_oids = [r[I_OID] for r in main_rows]
    max_pay = max((len(pay_by_oid.get(o, [])) for o in kept_oids), default=0)
    max_refund = max((len(ref_by_oid.get(o, [])) for o in kept_oids), default=0)
    print(f'  maxPay={max_pay}  maxRefund={max_refund}（按去重后集）')

    # 财年缺失校验：先扫一遍所有 biz_date 的财年，全部要在 SEASON 里
    missing = {}
    for row in main_rows:
        bd = row[idx['_biz_dt']]
        if bd is None:
            continue
        fy = fiscal_year(bd)
        if fy not in SEASON:
            missing[fy] = missing.get(fy, 0) + 1
    if missing:
        det = '，'.join(f'{k}（{v} 单）' for k, v in sorted(missing.items()))
        raise SystemExit(
            f'以下财年不在 SEASON 营业区间表中：{det}。\n'
            f'请在 export_rent_orders_fy.py 的 SEASON dict 补对应「财年→(营业起,营业止)」后重跑。')

    # ── 组装表头（与数据行同一处生成，杜绝错位）──
    headers = ['业务', '财年', '营/非', '财年序号', '运营日序号', '日序号', '月份',
               '创建日期', '创建时间', '支付次数', '支付合计', '退款次数', '退款合计',
               '订单结余', '订单状态', '最后退款日期', '最后退款时间']
    for k in range(1, max_pay + 1):
        headers += [f'【支付{k}】日期', f'【支付{k}】时间', f'【支付{k}】金额',
                    f'【支付{k}】支付方式', f'【支付{k}】支付账号']
    for k in range(1, max_refund + 1):
        headers += [f'【退款{k}】日期', f'【退款{k}】时间', f'【退款{k}】金额',
                    f'【退款{k}】退款方式']
    # 段4（订单状态/最后退款 已在段1）
    headers += ['超时费合计', '赔偿合计', '减免合计', '隐藏订单',
                '应分账金额', '实分账金额', '待分账金额', '业务', '门店',
                '客户名称', '电话', 'union id', '收款方式', '支付账号']
    # 段5
    headers += ['订单号', '业务日期', '业务时间', '结算日期', '结算时间',
                '支付总金额', '退款总金额', '订单结余', '店员姓名', '测试',
                '临时订单', '客户名称', '正/闭']

    # 金额列号（1-based）用于 number_format
    money_cols = set()
    for ci, h in enumerate(headers, start=1):
        if h in MONEY_FIXED or h in ('支付总金额', '退款总金额') \
           or h.endswith('】金额') or (h == '订单结余'):
            money_cols.add(ci)

    rows = []
    for row in main_rows:
        g = lambda n: row[idx[n]]
        oid = g('_oid')
        bd = g('_biz_dt')
        fy = fiscal_year(bd) if bd else ''
        s_start, s_end = (SEASON.get(fy, (None, None)))
        bdate = bd.date() if bd else None
        if bdate and s_start and s_start <= bdate <= s_end:
            ying = '营业'
            ops_day = (bdate - s_start).days + 1
        else:
            ying = '非营业'
            ops_day = None
        crt_d, crt_t = split_dt(g('_crt_dt'))
        lr_d, lr_t = split_dt(g('_last_refund_dt'))
        paid = g('支付合计') or 0
        refund = g('退款合计') or 0
        balance = g('订单结余') or 0
        test_flag = g('测试')
        rental_cnt = g('_rental_cnt') or 0
        status = derive_status(rental_cnt, g('_settled_cnt') or 0, g('_min_start'),
                               g('_any_end_null') or 0, paid, refund,
                               g('_rent_summary') or 0, g('_closed'))
        temp_order = '是' if (test_flag != '是' and balance > 0 and rental_cnt == 0) else ''
        zheng_bi = '关闭' if (paid == 0 and (g('_entertain') or 0) == 0) else '正常'

        seg1 = ['租赁', fy, ying, '', ops_day, g('日序号'), (bd.month if bd else None),
                crt_d, crt_t, g('支付次数') or 0, round(float(paid), 2),
                g('退款次数') or 0, round(float(refund), 2), round(float(balance), 2),
                status, lr_d, lr_t]

        seg2 = []
        plist = pay_by_oid.get(oid, [])
        for k in range(max_pay):
            if k < len(plist):
                pt, amt, pm, acct = plist[k]
                pd_, ptm = split_dt(pt)
                seg2 += [pd_, ptm, round(float(amt), 2) if amt is not None else None, pm, acct]
            else:
                seg2 += [None, None, None, None, None]

        seg3 = []
        rlist = ref_by_oid.get(oid, [])
        for k in range(max_refund):
            if k < len(rlist):
                rt, amt, rm = rlist[k]
                rd_, rtm = split_dt(rt)
                seg3 += [rd_, rtm, round(float(amt), 2) if amt is not None else None, rm]
            else:
                seg3 += [None, None, None, None]

        seg4 = [round(float(g('超时费合计') or 0), 2), round(float(g('赔偿合计') or 0), 2),
                round(float(g('减免合计') or 0), 2), g('隐藏订单'),
                round(float(g('应分账金额') or 0), 2), round(float(g('实分账金额') or 0), 2),
                round(float(g('待分账金额') or 0), 2), '租赁', g('门店'),
                g('客户名称'), g('电话'), g('union id'), g('收款方式'), g('支付账号')]

        seg5 = [g('订单号'), g('业务日期'), g('业务时间'), lr_d, lr_t,
                round(float(paid), 2), round(float(refund), 2), round(float(balance), 2),
                g('店员姓名'), test_flag, temp_order, g('客户名称'), zheng_bi]

        full = seg1 + seg2 + seg3 + seg4 + seg5
        assert len(full) == len(headers), f'列错位 row={len(full)} header={len(headers)}'
        rows.append(full)

    print(f'写 Excel: {out}  （{len(headers)} 列 × {len(rows)} 行）')
    wb = Workbook()
    write_sheet(wb.active, SHEET_TITLE, HEADER_COLOR, headers, rows)
    ws = wb.active
    for ci in money_cols:
        for r in range(2, ws.max_row + 1):
            ws.cell(row=r, column=ci).number_format = '0.00'
    wb.save(out)
    print(f'  完成。文件大小: {os.path.getsize(out) / 1024:.1f} KB')
    print(f'完成: {out}')


if __name__ == '__main__':
    main()
