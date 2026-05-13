#!/usr/bin/env python3
"""按店铺导出租赁订单 xlsx（订单汇总 / 订单明细 / 支付明细），含测试单标记 + 临时订单分类 + 对账标红。

用法：
    python3 export_rent_orders.py --shop 万龙体验中心 --start 2025-10-15 --end 2026-04-15
    python3 export_rent_orders.py --shop 渔阳 --start 2026-01-01 --end 2026-03-31 --out yuyang_q1.xlsx

环境要求（macOS）：
    brew install unixodbc msodbcsql18
    pip install pyodbc openpyxl
    export ODBCSYSINI=/opt/homebrew/etc   # 让 pyodbc 看到 ODBC Driver 18

数据库连接默认硬编码到生产，密码也在脚本里。换库时可用 --conn 覆盖或改 DEFAULT_CONN。
"""
import argparse
import os
import sys
from datetime import datetime, timedelta

import pyodbc
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding='utf-8')

DEFAULT_CONN = (
    'DRIVER={ODBC Driver 18 for SQL Server};SERVER=tcp:100.28.143.19,1433;'
    'DATABASE=snowmeet_new;UID=claude;PWD=abcd123!@#;'
    'Encrypt=yes;TrustServerCertificate=yes;Connection Timeout=30;'
)

# 退款条件：state=1 OR refund_id 非空（与 RentOrder.cs:519 旧代码一致）
REFUND_COND = "(pr.state = 1 OR (pr.refund_id IS NOT NULL AND pr.refund_id <> ''))"

# 已知店铺中文 → 文件名前缀（输出文件名默认 {prefix}_rent_orders_{start}_{end}.xlsx）
SHOP_PREFIX = {
    '万龙体验中心': 'wanlong',
    '万龙服务中心': 'wanlong_service',
    '渔阳': 'yuyang',
    '南山': 'nanshan',
    '怀北': 'huaibei',
    '崇礼旗舰店': 'chongli',
}

SUMMARY_SQL = f"""
SELECT
    o.code                                              AS 订单号,
    CAST(o.biz_date AS DATE)                            AS 业务日期,
    CONVERT(VARCHAR(8), o.biz_date, 108)                AS 业务时间,
    CAST(refund_agg.last_refund_date AS DATE)           AS 结算日期,
    CONVERT(VARCHAR(8), refund_agg.last_refund_date, 108) AS 结算时间,
    ISNULL(pay_agg.paid_amount, 0)                      AS 支付总金额,
    ISNULL(refund_agg.refund_amount, 0)                 AS 退款总金额,
    ISNULL(pay_agg.paid_amount, 0) - ISNULL(refund_agg.refund_amount, 0) AS 订单结余,
    s.name                                              AS 店员姓名,
    CASE WHEN ISNULL(pay_agg.paid_amount, 0) < 5
              OR (s.name IS NOT NULL AND s.name LIKE N'%苍%')
         THEN N'是' ELSE N'' END                        AS 测试
FROM [order] o
LEFT JOIN staff s ON s.id = o.staff_id
LEFT JOIN (
    SELECT order_id, SUM(amount) AS paid_amount
    FROM order_payment
    WHERE status = N'支付成功' AND valid = 1
    GROUP BY order_id
) pay_agg ON pay_agg.order_id = o.id
LEFT JOIN (
    SELECT pr.order_id,
           SUM(pr.amount)            AS refund_amount,
           MAX(pr.create_date)       AS last_refund_date
    FROM payment_refund pr
    WHERE {REFUND_COND}
    GROUP BY pr.order_id
) refund_agg ON refund_agg.order_id = o.id
WHERE o.shop = ?
  AND o.[type] = N'租赁'
  AND o.create_date >= ?
  AND o.create_date < ?
  AND o.valid = 1
ORDER BY o.create_date ASC, o.id ASC
"""

DETAIL_SQL = f"""
WITH base AS (
    SELECT
        o.id AS oid, o.code AS ocode, o.create_date AS ocrt,
        s.name AS sname,
        ISNULL(pa.paid_amount, 0) AS paid_amount,
        r.id AS rid, r.name AS rname, r.start_date, r.end_date,
        r.entertain, r.experience,
        ISNULL(SUM(CASE WHEN rd.charge_type = N'租金'   AND rd.valid = 1 THEN rd.amount END), 0) AS rent_total,
        ISNULL(SUM(CASE WHEN rd.charge_type = N'超时费' AND rd.valid = 1 THEN rd.amount END), 0) AS overtime,
        ISNULL(SUM(CASE WHEN rd.charge_type IN (N'赔偿金', N'损坏赔偿') AND rd.valid = 1 THEN rd.amount END), 0) AS damage
    FROM [order] o
    JOIN rental r ON r.order_id = o.id AND r.valid = 1
    LEFT JOIN rental_detail rd ON rd.rental_id = r.id
    LEFT JOIN staff s ON s.id = o.staff_id
    LEFT JOIN (
        SELECT order_id, SUM(amount) AS paid_amount
        FROM order_payment
        WHERE status = N'支付成功' AND valid = 1
        GROUP BY order_id
    ) pa ON pa.order_id = o.id
    WHERE o.shop = ?
      AND o.[type] = N'租赁'
      AND o.create_date >= ?
      AND o.create_date < ?
      AND o.valid = 1
    GROUP BY o.id, o.code, o.create_date, s.name, pa.paid_amount,
             r.id, r.name, r.start_date, r.end_date, r.entertain, r.experience
)
SELECT
    b.ocode                                              AS 订单号,
    b.rname                                              AS 租赁商品名称,
    CAST(b.start_date AS DATE)                           AS 起租日期,
    CONVERT(VARCHAR(8), b.start_date, 108)               AS 起租时间,
    CAST(b.end_date AS DATE)                             AS 退租日期,
    CONVERT(VARCHAR(8), b.end_date, 108)                 AS 退租时间,
    b.rent_total                                         AS 租金总额,
    CASE WHEN b.entertain  = 1 THEN N'是' ELSE N'' END   AS 是否招待,
    CASE WHEN b.experience = 1 THEN N'是' ELSE N'' END   AS 是否体验,
    CASE WHEN b.entertain = 1 OR b.experience = 1
         THEN 0 ELSE b.rent_total END                    AS 应付租金,
    ISNULL((
        SELECT SUM(d.amount) FROM discount d
        WHERE d.valid = 1 AND (
                 EXISTS (SELECT 1 FROM rental_detail rd
                         WHERE rd.id = d.sub_biz_id AND rd.rental_id = b.rid)
              OR (d.biz_type = N'租赁' AND d.biz_id = b.rid
                  AND NOT EXISTS (SELECT 1 FROM rental_detail rd2
                                  WHERE rd2.id = d.sub_biz_id AND rd2.rental_id = b.rid))
        )
    ), 0)                                                AS 减免金额,
    b.overtime                                           AS 超时费,
    b.damage                                             AS 损毁赔偿,
    (CASE WHEN b.entertain = 1 OR b.experience = 1 THEN 0 ELSE b.rent_total END)
        - ISNULL((
            SELECT SUM(d.amount) FROM discount d
            WHERE d.valid = 1 AND (
                     EXISTS (SELECT 1 FROM rental_detail rd
                             WHERE rd.id = d.sub_biz_id AND rd.rental_id = b.rid)
                  OR (d.biz_type = N'租赁' AND d.biz_id = b.rid
                      AND NOT EXISTS (SELECT 1 FROM rental_detail rd2
                                      WHERE rd2.id = d.sub_biz_id AND rd2.rental_id = b.rid))
            )
        ), 0)
        + b.overtime + b.damage                          AS 实付金额,
    CASE WHEN b.paid_amount < 5
              OR (b.sname IS NOT NULL AND b.sname LIKE N'%苍%')
         THEN N'是' ELSE N'' END                         AS 测试
FROM base b
ORDER BY b.ocrt ASC, b.oid ASC, b.rid ASC
"""

PAYMENT_SQL = f"""
SELECT
    o.code                                                  AS 订单号,
    op.pay_method                                           AS 支付方式,
    CASE WHEN op.pay_method = N'微信支付' THEN wk.mch_id ELSE NULL END AS mch_id,
    op.amount                                               AS 支付金额,
    ISNULL(rp.refund_amount, 0)                             AS 退款金额,
    op.amount - ISNULL(rp.refund_amount, 0)                 AS 结余金额,
    CASE WHEN ISNULL(pa.paid_amount, 0) < 5
              OR (s.name IS NOT NULL AND s.name LIKE N'%苍%')
         THEN N'是' ELSE N'' END                            AS 测试
FROM [order] o
JOIN order_payment op ON op.order_id = o.id
LEFT JOIN wepay_key wk ON wk.id = op.mch_id
LEFT JOIN staff s ON s.id = o.staff_id
LEFT JOIN (
    SELECT order_id, SUM(amount) AS paid_amount
    FROM order_payment
    WHERE status = N'支付成功' AND valid = 1
    GROUP BY order_id
) pa ON pa.order_id = o.id
LEFT JOIN (
    SELECT pr.payment_id, SUM(pr.amount) AS refund_amount
    FROM payment_refund pr
    WHERE {REFUND_COND}
    GROUP BY pr.payment_id
) rp ON rp.payment_id = op.id
WHERE o.shop = ?
  AND o.[type] = N'租赁'
  AND o.create_date >= ?
  AND o.create_date < ?
  AND o.valid = 1
  AND op.status = N'支付成功'
  AND op.valid = 1
ORDER BY o.create_date ASC, o.id ASC, op.id ASC
"""


def write_sheet(ws, title, header_color, headers, rows):
    ws.title = title
    ws.append(headers)
    fill = PatternFill('solid', fgColor=header_color)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    for row in rows:
        ws.append(list(row))
    ws.freeze_panes = 'A2'
    for c in range(1, len(headers) + 1):
        col = get_column_letter(c)
        max_len = sum(2 if ord(ch) > 127 else 1 for ch in str(headers[c - 1]))
        for r in range(2, min(len(rows), 200) + 2):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            s = v.strftime('%Y-%m-%d %H:%M') if isinstance(v, datetime) else str(v)
            wlen = sum(2 if ord(ch) > 127 else 1 for ch in s)
            if wlen > max_len:
                max_len = wlen
        ws.column_dimensions[col].width = min(max_len + 2, 36)


def post_process(out_path):
    """对账后处理：
    1. 订单汇总新增「临时订单」列：非测试订单中，订单结余 > 0 且订单明细无非测试 rental 行 → '是'
    2. 订单结余 vs 订单明细实付合计差额 ≥ 0.01 且非临时订单 → 订单号 cell 浅红底 + 深红字
    """
    wb = load_workbook(out_path)
    ws_sum = wb['订单汇总']
    ws_det = wb['订单明细']
    sum_h = [c.value for c in ws_sum[1]]
    det_h = [c.value for c in ws_det[1]]
    SUM_CODE = sum_h.index('订单号') + 1
    SUM_BALANCE = sum_h.index('订单结余') + 1
    SUM_TEST = sum_h.index('测试') + 1
    DET_CODE = det_h.index('订单号') + 1
    DET_PAID = det_h.index('实付金额') + 1
    DET_TEST = det_h.index('测试') + 1

    # 按订单号聚合订单明细的非测试 rental 实付合计
    det_paid_sum = {}
    for r in range(2, ws_det.max_row + 1):
        if ws_det.cell(row=r, column=DET_TEST).value == '是':
            continue
        code = ws_det.cell(row=r, column=DET_CODE).value
        paid = ws_det.cell(row=r, column=DET_PAID).value or 0
        det_paid_sum[code] = det_paid_sum.get(code, 0) + paid

    # 加「临时订单」列
    tmp_col = ws_sum.max_column + 1
    head_cell = ws_sum.cell(row=1, column=1)
    new_head = ws_sum.cell(row=1, column=tmp_col, value='临时订单')
    if head_cell.has_style:
        from copy import copy
        new_head.font = copy(head_cell.font)
        new_head.fill = copy(head_cell.fill)
        new_head.alignment = copy(head_cell.alignment)
        new_head.border = copy(head_cell.border)

    red_fill = PatternFill('solid', fgColor='FFC7CE')
    red_color = 'C00000'

    tmp_count = mismatch_count = 0
    for r in range(2, ws_sum.max_row + 1):
        if ws_sum.cell(row=r, column=SUM_TEST).value == '是':
            continue
        code = ws_sum.cell(row=r, column=SUM_CODE).value
        balance = ws_sum.cell(row=r, column=SUM_BALANCE).value or 0
        det_total = det_paid_sum.get(code)

        # 临时订单：结余>0 但订单明细里 0 条非测试 rental 行
        if det_total is None and balance > 0:
            ws_sum.cell(row=r, column=tmp_col, value='是')
            tmp_count += 1
            continue

        # 对账差异（剩余 rental 存在但金额不等）
        if det_total is not None and abs(round(balance - det_total, 2)) >= 0.01:
            cell = ws_sum.cell(row=r, column=SUM_CODE)
            cell.fill = red_fill
            old = cell.font
            cell.font = Font(name=old.name, size=old.size, bold=old.bold,
                             italic=old.italic, color=red_color)
            mismatch_count += 1

    wb.save(out_path)
    return tmp_count, mismatch_count


def default_out_name(shop, start, end):
    prefix = SHOP_PREFIX.get(shop, shop)
    return f'{prefix}_rent_orders_{start}_{end}.xlsx'


def parse_args():
    p = argparse.ArgumentParser(
        description='导出某店铺某时间段的租赁订单 xlsx（含 3 个 sheet + 对账标记）',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    p.add_argument('--shop', required=True, help='店铺名（DB order.shop 字段值，如「万龙体验中心」）')
    p.add_argument('--start', required=True, help='起始日期（inclusive），格式 YYYY-MM-DD')
    p.add_argument('--end', required=True, help='截止日期（inclusive），格式 YYYY-MM-DD')
    p.add_argument('--out', default=None, help='输出 xlsx 路径，默认按 {shop_prefix}_rent_orders_{start}_{end}.xlsx 生成在当前目录')
    p.add_argument('--conn', default=DEFAULT_CONN, help='ODBC 连接字符串，默认连生产')
    p.add_argument('--no-postprocess', action='store_true', help='跳过对账后处理（不加「临时订单」列、不标红）')
    return p.parse_args()


def main():
    args = parse_args()
    start = args.start
    end_excl = (datetime.strptime(args.end, '%Y-%m-%d') + timedelta(days=1)).strftime('%Y-%m-%d')
    out = args.out or default_out_name(args.shop, args.start, args.end)
    out = os.path.abspath(out)

    print('连数据库 ...')
    cn = pyodbc.connect(args.conn)
    cur = cn.cursor()

    print(f'查汇总（{args.shop} / {start} ~ {end_excl}） ...')
    cur.execute(SUMMARY_SQL, args.shop, start, end_excl)
    summary_cols = [c[0] for c in cur.description]
    summary_rows = cur.fetchall()
    print(f'  汇总: {len(summary_rows)} 行')

    print('查明细 ...')
    cur.execute(DETAIL_SQL, args.shop, start, end_excl)
    detail_cols = [c[0] for c in cur.description]
    detail_rows = cur.fetchall()
    print(f'  明细: {len(detail_rows)} 行')

    print('查支付明细 ...')
    cur.execute(PAYMENT_SQL, args.shop, start, end_excl)
    pay_cols = [c[0] for c in cur.description]
    pay_rows = cur.fetchall()
    print(f'  支付明细: {len(pay_rows)} 行')

    cn.close()

    print(f'写 Excel: {out}')
    wb = Workbook()
    write_sheet(wb.active,        '订单汇总', '1F4E78', summary_cols, summary_rows)
    write_sheet(wb.create_sheet(), '订单明细', '2E7D32', detail_cols, detail_rows)
    write_sheet(wb.create_sheet(), '支付明细', 'B7791F', pay_cols,    pay_rows)
    wb.save(out)
    size_kb = os.path.getsize(out) / 1024
    print(f'  完成。文件大小: {size_kb:.1f} KB')

    if not args.no_postprocess:
        print('对账后处理 ...')
        tmp_n, mis_n = post_process(out)
        print(f'  临时订单标记: {tmp_n} 行')
        print(f'  订单结余 ≠ 明细实付（订单号标红）: {mis_n} 行')

    print(f'完成: {out}')


if __name__ == '__main__':
    main()
