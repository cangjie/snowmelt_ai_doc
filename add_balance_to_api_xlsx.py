"""一次性补列：把 wanlong_rent_orders_2025-10-15_2026-04-15.xlsx「订单汇总」sheet 的「订单结余」
按订单号查表写入 wanlong_rent_orders_api_2025-10-15_2026-04-15.xlsx 的「订单」sheet 末尾。"""
import os
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding='utf-8')

BASE = os.path.dirname(os.path.abspath(__file__))
SRC = os.path.join(BASE, "wanlong_rent_orders_2025-10-15_2026-04-15.xlsx")
TGT = os.path.join(BASE, "wanlong_rent_orders_api_2025-10-15_2026-04-15.xlsx")
SRC_SHEET = "订单汇总"
TGT_SHEET = "订单"
NEW_HEADER = "订单结余"
KEY_HEADER = "订单号"


def visual_len(s):
    return sum(2 if ord(ch) > 127 else 1 for ch in str(s))


def build_balance_map():
    wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
    ws = wb[SRC_SHEET]
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    try:
        key_idx = header.index(KEY_HEADER)
        bal_idx = header.index(NEW_HEADER)
    except ValueError as e:
        raise SystemExit(f"源表头缺字段：{e}；实际表头={header}")
    m = {}
    for row in rows:
        if row is None:
            continue
        code = row[key_idx]
        if code is None:
            continue
        m[code] = row[bal_idx]
    wb.close()
    return m


def main():
    bal_map = build_balance_map()
    print(f"源 dict size: {len(bal_map)}")

    wb = openpyxl.load_workbook(TGT)
    ws = wb[TGT_SHEET]

    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    if NEW_HEADER in headers:
        new_col = headers.index(NEW_HEADER) + 1
        print(f"已存在「{NEW_HEADER}」列（第 {new_col} 列），将覆盖写入")
    else:
        new_col = ws.max_column + 1

    header_cell = ws.cell(row=1, column=new_col, value=NEW_HEADER)
    header_cell.font = Font(bold=True, color="FFFFFF")
    header_cell.fill = PatternFill("solid", fgColor="1F4E78")
    header_cell.alignment = Alignment(horizontal="center", vertical="center")

    NUM_FMT = "0.00"
    miss = []
    last_row = ws.max_row
    for r in range(2, last_row + 1):
        code = ws.cell(row=r, column=1).value
        if code is None:
            continue
        # 总计租金列 (D) 一并清浮点尾巴，避免 General 格式触发科学计数法
        rent_cell = ws.cell(row=r, column=4)
        if isinstance(rent_cell.value, (int, float)):
            rent_cell.value = round(float(rent_cell.value), 2)
            rent_cell.number_format = NUM_FMT
        if code in bal_map:
            v = bal_map[code]
            if isinstance(v, (int, float)):
                v = round(float(v), 2)
            bal_cell = ws.cell(row=r, column=new_col, value=v)
            if isinstance(v, (int, float)):
                bal_cell.number_format = NUM_FMT
        else:
            miss.append((r, code))

    max_len = visual_len(NEW_HEADER)
    for r in range(2, min(last_row, 200) + 1):
        v = ws.cell(row=r, column=new_col).value
        if v is None:
            continue
        wlen = visual_len(v)
        if wlen > max_len:
            max_len = wlen
    ws.column_dimensions[get_column_letter(new_col)].width = min(max_len + 2, 36)

    wb.save(TGT)
    size_kb = os.path.getsize(TGT) / 1024

    print(f"目标行数: {last_row - 1}")
    print(f"未命中: {len(miss)}")
    for r, code in miss[:5]:
        print(f"  - 行 {r} 订单号 {code}")
    print(f"写入 {TGT}\n文件大小: {size_kb:.1f} KB")


if __name__ == "__main__":
    main()
