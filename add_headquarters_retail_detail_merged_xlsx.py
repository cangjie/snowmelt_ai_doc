"""一次性产物：总部零售财年报表按「七色米订单号」合并展示商品明细。

规则完全沿用崇礼版 add_chongli_retail_detail_merged_xlsx.py（亦即南山版
add_retail_detail_merged_xlsx.py），仅换输入/输出文件。

匹配键：
    年度零售「七色米订单号」 == 销售明细单1「单据编号」（均为 XSDxxxx 格式）

明细源说明（与崇礼一致，单一统一源）：
    明细源是 all_销售单列表.xls —— 七色米全店全量导出（含崇礼/万龙/南山/
    总部等所有门店）。因此本脚本只做「报表订单 → 明细」的正向匹配；
    **不做**「明细记录是否存在于本报表」的反向核对（某条明细不在总部，
    可能属于其他门店，属正常）。
    红色 FF9999 仍保留：表示「本报表内非关闭单未匹配到明细」的正向异常。

输出形态（与崇礼/南山一致）：
    [1] 另存独立备份 headquarters_retail_orders_fy_with_detail.xlsx（单 sheet）
    [2] 幂等放入主报表（已存在「年度零售明细」则删重建，其余 sheet 不动）
    新 sheet「年度零售明细」：原订单级列 + 10 明细列 + 末列「Σ明细总额−订单结余」
    一单多商品 → 纵向展开 N 行，订单级列合并跨 N 行；
    无七色米号 / 未命中明细 → 单行，明细 10 列留空。

用法（确保产物未被 Excel 打开；源报表被打开不影响，只读加载）：
    python3 add_headquarters_retail_detail_merged_xlsx.py
"""
import os
import sys
from collections import defaultdict

import xlrd  # noqa: E402  读 .xls
import openpyxl  # noqa: E402  读/写 .xlsx
from openpyxl.styles import Font, PatternFill, Alignment  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

sys.stdout.reconfigure(encoding="utf-8")

BASE = os.path.dirname(os.path.abspath(__file__))
SRC_XLS = os.path.join(BASE, "all_销售单列表.xls")
SRC_XLSX = os.path.join(
    BASE, "headquarters_retail_orders_fy_2025-05-01_2026-04-30.xlsx")
OUT_XLSX = os.path.join(
    BASE, "headquarters_retail_orders_fy_with_detail.xlsx")

DETAIL_SHEET = "销售明细单1"
SRC_SHEET = "年度零售"
OUT_SHEET = "年度零售明细"
KEY_HEADER = "七色米订单号"          # 年度零售 里的匹配列
DOC_HEADER = "单据编号"              # 销售明细单1 里的匹配列
HEADER_FILL = "1F4E78"
MULTI_FILL = "EAF2FB"  # >1 条明细的订单块底色（浅蓝）
DIFF_FILL = "FCE4EC"   # 该单Σ明细总额 与 订单结余 有差额 → 淡粉
NODETAIL_FILL = "FF9999"  # 非关闭无明细（对不上）单 → 标红
AGG_HEADER = "Σ明细总额−订单结余"  # 末列：每单(明细总额合计 − 订单结余) 有符号差额
BAL_HEADER = "订单结余"        # 差额基准：与该单Σ明细总额比对
DIFF_TOL = 0.01               # 差额阈值（¥）
CLOSED_HEADER = "正/闭"        # 关闭单不匹配明细
CLOSED_VALUE = "关闭"          # 该列取值仅 关闭/正常
# 总部首次合并：暂无用户指定剔除测试单。沿用既定口径——微额 ¥0.0x/¥0 无号单
# 属测试单，实额缺号单保持标红，剔除范围待用户单点确认后再补入本集合。
EXCLUDE_CODES = set()
CODE_HEADER = "订单号"

# 明细 10 字段：(销售明细单1 表头名, 是否数值列)
DETAIL_FIELDS = [
    ("商品编号", False),
    ("商品名称", False),
    ("商品分类", False),
    ("规格", False),
    ("属性", False),
    ("数量", True),
    ("单价", True),
    ("折扣", True),
    ("折后单价", True),
    ("总额", True),
]
# 在新 sheet 里强制 0.00 显示的明细列（按 DETAIL_FIELDS 的 0 基序号）
MONEY_DETAIL_IDX = [6, 8, 9]  # 单价 / 折后单价 / 总额


def visual_len(s):
    return sum(2 if ord(ch) > 127 else 1 for ch in str(s))


def _txt(v):
    """文本字段：空→None，否则 str 去空白（防 xlrd 把纯数字编号读成 float）。"""
    if v is None:
        return None
    if isinstance(v, float):
        # 商品编号偶发被识别为数字时还原成无小数无科学计数法的串
        s = ("%d" % v) if v == int(v) else repr(v)
    else:
        s = str(v)
    s = s.strip()
    return s or None


def _num(v):
    """数值字段：空字符串/None→None，数字原样保留。"""
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return v
    s = str(v).strip()
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return s  # 兜底：保留原文（如 '-'）


def load_detail_map():
    """读销售明细单1 → {单据编号: [(10字段值...), ...]}，保留 sheet 行序。"""
    wb = xlrd.open_workbook(SRC_XLS)
    if DETAIL_SHEET not in wb.sheet_names():
        raise SystemExit(f"sheet '{DETAIL_SHEET}' 不存在，实际: {wb.sheet_names()}")
    sh = wb.sheet_by_name(DETAIL_SHEET)
    hdr = [str(sh.cell_value(0, c)).strip() for c in range(sh.ncols)]

    def col_of(name):
        if name not in hdr:
            raise SystemExit(f"销售明细单1 缺表头「{name}」: {hdr}")
        return hdr.index(name)

    doc_c = col_of(DOC_HEADER)
    field_cols = [(col_of(n), is_num) for n, is_num in DETAIL_FIELDS]

    dmap = defaultdict(list)
    n_rows = 0
    for r in range(1, sh.nrows):
        doc = str(sh.cell_value(r, doc_c)).strip()
        if not doc:
            continue
        rec = []
        for c, is_num in field_cols:
            raw = sh.cell_value(r, c)
            rec.append(_num(raw) if is_num else _txt(raw))
        dmap[doc].append(tuple(rec))
        n_rows += 1
    return dmap, len(dmap), n_rows


def load_source():
    """读年度零售 → (headers, data_rows[list[tuple]], key_idx0)。"""
    wb = openpyxl.load_workbook(SRC_XLSX, read_only=True, data_only=True)
    if SRC_SHEET not in wb.sheetnames:
        raise SystemExit(f"sheet '{SRC_SHEET}' 不存在，实际: {wb.sheetnames}")
    ws = wb[SRC_SHEET]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    headers = list(rows[0])
    if KEY_HEADER not in headers:
        raise SystemExit(f"年度零售 缺表头「{KEY_HEADER}」: {headers}")
    key_idx0 = headers.index(KEY_HEADER)
    data = [r for r in rows[1:] if any(v is not None for v in r)]
    return headers, data, key_idx0


def build(headers, data, key_idx0, dmap):
    """展开行 + 合并区间 + 末列差额 + 染粉区间 + 标红区间。
    返回 (out_rows, merge_spans, pink_spans, red_spans, stats)。
    关闭单（正/闭=关闭）/ 测试单（EXCLUDE_CODES）整单从本 sheet 删除。
    末列每单仅 anchor 行有值（订单级，随订单列一起合并）。
    pink_spans: 匹配单中 |Σ明细总额 − 订单结余| > DIFF_TOL 的 (start,end)。
    red_spans: 非关闭非剔除但无明细（对不上）的单行 (ridx,ridx)，整行标红。"""
    n_orig = len(headers)
    tot_off = next(i for i, (nm, _) in enumerate(DETAIL_FIELDS) if nm == "总额")
    bal_idx0 = headers.index(BAL_HEADER)  # 第一个「订单结余」（段1）
    closed_idx0 = headers.index(CLOSED_HEADER)
    code_idx0 = headers.index(CODE_HEADER)
    out_rows = []
    merge_spans = []
    pink_spans = []
    red_spans = []
    st = {"retail": len(data), "mi7": 0, "matched": 0,
          "no_detail": 0, "merged_orders": 0, "diff": 0,
          "closed": 0, "excluded": 0}

    for r in data:
        # 用户指定剔除的测试单：整单删除
        if r[code_idx0] in EXCLUDE_CODES:
            st["excluded"] += 1
            continue
        # 关闭单整单从「年度零售明细」删除（不输出任何行）
        if str(r[closed_idx0]).strip() == CLOSED_VALUE:
            st["closed"] += 1
            continue
        mi7 = r[key_idx0]
        mi7 = str(mi7).strip() if mi7 is not None else ""
        if mi7:
            st["mi7"] += 1
        dets = dmap.get(mi7) if mi7 else None
        if dets:
            st["matched"] += 1
            start_row = len(out_rows) + 2  # +1 表头, +1 转 1 基
            end_row = start_row + len(dets) - 1
            order_total = round(sum(d[tot_off] for d in dets
                                    if isinstance(d[tot_off], (int, float))), 2)
            bal = r[bal_idx0]
            diff_val = (round(order_total - bal, 2)
                        if isinstance(bal, (int, float)) else None)
            for k, det in enumerate(dets):
                if k == 0:
                    out_rows.append(list(r) + list(det) + [diff_val])
                else:
                    out_rows.append([None] * n_orig + list(det) + [None])
            if len(dets) >= 2:
                merge_spans.append((start_row, end_row))
                st["merged_orders"] += 1
            if diff_val is not None and abs(diff_val) > DIFF_TOL:
                pink_spans.append((start_row, end_row))
                st["diff"] += 1
        else:
            # 非关闭、未剔除但无明细 → 对不上，整行标红
            st["no_detail"] += 1
            ridx = len(out_rows) + 2  # +1 表头, +1 转 1 基
            out_rows.append(
                list(r) + [None] * len(DETAIL_FIELDS) + [None])
            red_spans.append((ridx, ridx))
    return out_rows, merge_spans, pink_spans, red_spans, st


def render_sheet(ws, out_headers, out_rows, merge_spans, pink_spans,
                 red_spans, n_orig):
    """把表头/行/合并/底色/数字格式/列宽渲染进给定 worksheet。"""
    total_cols = len(out_headers)
    ws.append(out_headers)
    fill = PatternFill("solid", fgColor=HEADER_FILL)
    for c in range(1, total_cols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in out_rows:
        ws.append(row)
    ws.freeze_panes = "A2"

    # 合并订单级 n_orig 列 + 末列「Σ明细总额−订单结余」（订单级，一起合并）
    center = Alignment(vertical="center", wrap_text=False)
    order_cols = list(range(1, n_orig + 1)) + [total_cols]
    for s, e in merge_spans:
        for col in order_cols:
            ws.merge_cells(start_row=s, start_column=col,
                           end_row=e, end_column=col)
            ws.cell(row=s, column=col).alignment = center

    # 底色优先级：浅蓝(>1明细) < 淡粉(差额) < 红(无明细对不上)
    multi_fill = PatternFill("solid", fgColor=MULTI_FILL)
    diff_fill = PatternFill("solid", fgColor=DIFF_FILL)
    red_fill = PatternFill("solid", fgColor=NODETAIL_FILL)
    for s, e in merge_spans:
        for rr in range(s, e + 1):
            for col in range(1, total_cols + 1):
                ws.cell(row=rr, column=col).fill = multi_fill
    for s, e in pink_spans:
        for rr in range(s, e + 1):
            for col in range(1, total_cols + 1):
                ws.cell(row=rr, column=col).fill = diff_fill
    for s, e in red_spans:
        for rr in range(s, e + 1):
            for col in range(1, total_cols + 1):
                ws.cell(row=rr, column=col).fill = red_fill

    # 金额列锁 0.00（明细 单价/折后单价/总额 + 末列 Σ明细总额−订单结余）
    last_row = len(out_rows) + 1
    money_cols = [n_orig + 1 + off for off in MONEY_DETAIL_IDX] + [total_cols]
    for col in money_cols:
        for rr in range(2, last_row + 1):
            cell = ws.cell(row=rr, column=col)
            if isinstance(cell.value, (int, float)):
                cell.number_format = "0.00"

    # 自适应列宽（仿 write_sheet：表头 + 抽样数据，cap 36）
    sample_to = min(len(out_rows), 300) + 2
    for c in range(1, total_cols + 1):
        max_len = visual_len(out_headers[c - 1])
        for rr in range(2, sample_to):
            v = ws.cell(row=rr, column=c).value
            if v is None:
                continue
            wlen = visual_len(v)
            if wlen > max_len:
                max_len = wlen
        ws.column_dimensions[get_column_letter(c)].width = min(max_len + 2, 36)


def main():
    for p in (SRC_XLS, SRC_XLSX):
        if not os.path.exists(p):
            raise SystemExit(f"输入不存在: {p}")

    dmap, n_docs, n_det_rows = load_detail_map()
    print(f"销售明细单1: {n_docs} 单据 / {n_det_rows} 明细行（all 全店全量）")

    headers, data, key_idx0 = load_source()
    print(f"年度零售: {len(headers)} 列 / {len(data)} 数据行（七色米列 idx0={key_idx0}）")

    out_headers = list(headers) + [n for n, _ in DETAIL_FIELDS] + [AGG_HEADER]
    out_rows, merge_spans, pink_spans, red_spans, st = build(
        headers, data, key_idx0, dmap)

    # === 统计 + 断言 ===
    print("\n=== JOIN 统计 ===")
    print(f"  零售数据行 : {st['retail']}")
    print(f"  有七色米号 : {st['mi7']}")
    print(f"  匹配上明细 : {st['matched']}")
    print(f"  无明细单行 : {st['no_detail']}  (非关闭·对不上 → 整行标红)")
    print(f"  关闭单     : {st['closed']}  (已整单从「{OUT_SHEET}」删除，不输出)")
    print(f"  剔除单     : {st['excluded']}  "
          f"({'/'.join(sorted(EXCLUDE_CODES)) or '无'})")
    print(f"  需合并单数 : {st['merged_orders']}  (span>=2)")
    print(f"  差额单数   : {st['diff']} / {st['matched']}  "
          f"(末列 |Σ明细总额−订单结余| > ¥{DIFF_TOL}，淡粉标注)")
    print(f"  输出数据行 : {len(out_rows)}")
    span_sum = sum(e - s + 1 for s, e in merge_spans)
    single_or_one = len(out_rows) - span_sum
    print(f"  Σ合并跨度  : {span_sum}  (+ 非合并 {single_or_one} = {len(out_rows)})")
    assert st["retail"] == len(data)
    assert (st["matched"] + st["no_detail"] + st["closed"]
            + st["excluded"] == st["retail"])

    n_orig = len(headers)

    # === 1) 另存独立文件（备份留存）===
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = OUT_SHEET
    render_sheet(ws, out_headers, out_rows, merge_spans, pink_spans,
                 red_spans, n_orig)
    wb.save(OUT_XLSX)
    print(f"\n[1/2] 另存独立文件 {OUT_XLSX}  "
          f"({os.path.getsize(OUT_XLSX) / 1024:.1f} KB)")

    # === 2) 直接放入主报表（幂等：已存在则删重建；其余 sheet 不动）===
    if not os.path.exists(SRC_XLSX):
        raise SystemExit(f"主报表不存在: {SRC_XLSX}")
    mwb = openpyxl.load_workbook(SRC_XLSX)
    before = list(mwb.sheetnames)
    if OUT_SHEET in mwb.sheetnames:
        print(f"  主报表已存在「{OUT_SHEET}」，删除重建")
        del mwb[OUT_SHEET]
    mws = mwb.create_sheet(OUT_SHEET)
    render_sheet(mws, out_headers, out_rows, merge_spans, pink_spans,
                 red_spans, n_orig)
    mwb.save(SRC_XLSX)
    print(f"[2/2] 已放入主报表 {SRC_XLSX}  "
          f"({os.path.getsize(SRC_XLSX) / 1024:.1f} KB)")
    print(f"  原 sheets: {before}")
    print(f"  现 sheets: {list(mwb.sheetnames)}  "
          f"({len(out_headers)} 列 × {len(out_rows)} 数据行)")

    # === 抽检（小数据集可能无合并单，空集保护）===
    if merge_spans:
        print("\n=== 抽检：合并跨度最大 / 一个 2 明细单 ===")
        by_span = sorted(merge_spans, key=lambda se: se[1] - se[0],
                         reverse=True)
        picks = [("最大", by_span[0])]
        two = next((se for se in merge_spans if se[1] - se[0] == 1), None)
        if two:
            picks.append(("2明细", two))
        for label, (s, e) in picks:
            code = ws.cell(row=s, column=headers.index("订单号") + 1).value
            mi7 = ws.cell(row=s, column=key_idx0 + 1).value
            print(f"  [{label}] 行{s}~{e} ({e - s + 1}件) "
                  f"订单号={code} 七色米={mi7}")
            for rr in range(s, e + 1):
                vals = [ws.cell(row=rr, column=n_orig + 1 + k).value
                        for k in range(len(DETAIL_FIELDS))]
                print(f"    R{rr}: {vals}")
    else:
        print("\n=== 抽检：无合并单（所有匹配单均单明细行）===")

    # === 软核对：Σ明细总额 vs 年度零售「销售额合计」===
    if "销售额合计" in headers:
        sa_idx0 = headers.index("销售额合计")
        cl_idx0 = headers.index(CLOSED_HEADER)
        tot_idx0 = next(i for i, (n, _) in enumerate(DETAIL_FIELDS)
                        if n == "总额")
        mism = 0
        for r in data:
            if str(r[cl_idx0]).strip() == CLOSED_VALUE:
                continue  # 关闭单已从 sheet 删除，软核对一并跳过
            mi7 = r[key_idx0]
            mi7 = str(mi7).strip() if mi7 is not None else ""
            dets = dmap.get(mi7) if mi7 else None
            if not dets:
                continue
            try:
                ssum = round(sum(float(d[tot_idx0]) for d in dets
                                 if isinstance(d[tot_idx0], (int, float))), 2)
                sa = round(float(r[sa_idx0]), 2) if r[sa_idx0] is not None else 0.0
            except (TypeError, ValueError):
                continue
            if abs(ssum - sa) > 0.01:
                mism += 1
        print(f"\n软核对: 匹配单 Σ明细总额 vs 销售额合计 差>¥0.01 的单数 = "
              f"{mism}/{st['matched']}  (七色米侧总额 与 DB deal_price 口径"
              f"可能不同，属预期，仅提示)")


if __name__ == "__main__":
    main()
