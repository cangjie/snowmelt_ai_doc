"""导出 2025-10-15 ~ 2026-04-15 租赁订单 汇总 + 明细 到 Excel"""
import pyodbc, sys, os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding='utf-8')

CONN = ('DRIVER={ODBC Driver 18 for SQL Server};SERVER=tcp:100.28.143.19,1433;'
        'DATABASE=snowmeet_new;UID=claude;PWD=abcd123!@#;'
        'Encrypt=yes;TrustServerCertificate=yes;Connection Timeout=30;')
START = '2025-10-15'
END_EXCL = '2026-04-16'  # exclusive
OUT = r'D:\snowmeet\snowmeet_ai_doc\rent_orders_2025-10-15_2026-04-15.xlsx'

SUMMARY_SQL = f"""
SELECT
    o.id                         AS 订单ID,
    o.code                       AS 订单号,
    o.shop                       AS 店铺,
    o.sub_type                   AS 子类型,
    o.contact_name               AS 联系人,
    o.contact_num                AS 联系手机,
    o.contact_gender             AS 联系人性别,
    o.member_id                  AS 会员ID,
    m.real_name                  AS 会员姓名,
    s.name                       AS 业务员,
    o.biz_date                   AS 业务日期,
    o.create_date                AS 创建时间,
    o.close_date                 AS 关闭时间,
    o.total_amount               AS 订单总额,
    o.paying_amount              AS 待付金额,
    ISNULL(pay_agg.paid_amount, 0)        AS 实付金额,
    ISNULL(pay_agg.pay_methods, '')       AS 支付方式,
    ISNULL(rent_agg.package_count, 0)     AS 套餐数,
    ISNULL(rent_agg.guaranty_total, 0)    AS 押金合计,
    ISNULL(item_agg.item_count, 0)        AS 装备数,
    CASE
        WHEN o.recepting = 1 THEN N'接待中'
        WHEN o.closed = 1 THEN N'已关闭'
        WHEN o.entertain = 1 THEN N'招待'
        ELSE N'已下单' END                AS 状态,
    o.entertain                  AS 招待标记,
    o.pay_option                 AS 付款方式选项,
    o.is_test                    AS 测试单,
    o.memo                       AS 备注
FROM [order] o
LEFT JOIN staff s  ON s.id = o.staff_id
LEFT JOIN member m ON m.id = o.member_id
LEFT JOIN (
    SELECT order_id,
           SUM(amount) AS paid_amount,
           STUFF((SELECT DISTINCT ',' + p2.pay_method
                  FROM order_payment p2
                  WHERE p2.order_id = p1.order_id AND p2.status = N'支付成功' AND p2.valid = 1
                  FOR XML PATH('')), 1, 1, '') AS pay_methods
    FROM order_payment p1
    WHERE p1.status = N'支付成功' AND p1.valid = 1
    GROUP BY order_id
) pay_agg ON pay_agg.order_id = o.id
LEFT JOIN (
    SELECT order_id,
           COUNT(*) AS package_count,
           SUM(CASE WHEN noGuaranty=0 THEN ISNULL(guaranty,0) - ISNULL(guaranty_discount,0) ELSE 0 END) AS guaranty_total
    FROM rental
    WHERE valid = 1
    GROUP BY order_id
) rent_agg ON rent_agg.order_id = o.id
LEFT JOIN (
    SELECT r.order_id, COUNT(*) AS item_count
    FROM rent_item ri
    JOIN rental r ON r.id = ri.rental_id
    WHERE ri.valid = 1 AND r.valid = 1 AND ri.noNeed = 0
    GROUP BY r.order_id
) item_agg ON item_agg.order_id = o.id
WHERE o.[type] = N'租赁'
  AND o.create_date >= ?
  AND o.create_date < ?
  AND o.valid = 1
ORDER BY o.create_date ASC, o.id ASC
"""

DETAIL_SQL = f"""
SELECT
    o.id                                  AS 订单ID,
    o.code                                AS 订单号,
    o.shop                                AS 店铺,
    o.biz_date                            AS 业务日期,
    o.contact_name                        AS 联系人,
    o.contact_num                         AS 联系手机,
    r.id                                  AS 套餐ID,
    r.name                                AS 套餐名,
    r.package_id                          AS 套餐编号,
    r.category_id                         AS 品类ID,
    r.start_date                          AS 起租时间,
    r.end_date                            AS 应还时间,
    r.guaranty                            AS 押金,
    r.guaranty_discount                   AS 押金减免,
    CASE WHEN r.noGuaranty=0 THEN ISNULL(r.guaranty,0) - ISNULL(r.guaranty_discount,0) ELSE 0 END AS 押金净额,
    r.noGuaranty                          AS 免押金,
    r.pick_type                           AS 取货方式,
    r.settled                             AS 已结算,
    r.experience                          AS 体验,
    r.entertain                           AS 套餐招待,
    r.memo                                AS 套餐备注,
    ri.id                                 AS 装备ID,
    ri.name                               AS 装备名称,
    ri.class_name                         AS 装备类别,
    ri.code                               AS 装备编码,
    ri.noCode                             AS 无编码,
    ri.noNeed                             AS 不需要,
    ri.is_associate                       AS 附属项,
    ri.atOnce                             AS 立即租赁,
    ri.pick_time                          AS 取出时间,
    ri.return_time                        AS 归还时间,
    ri.pick_type                          AS 装备取货方式,
    ri.memo                               AS 装备备注
FROM [order] o
JOIN rental r       ON r.order_id = o.id AND r.valid = 1
LEFT JOIN rent_item ri ON ri.rental_id = r.id AND ri.valid = 1
WHERE o.[type] = N'租赁'
  AND o.create_date >= ?
  AND o.create_date < ?
  AND o.valid = 1
ORDER BY o.create_date ASC, o.id ASC, r.id ASC, ri.id ASC
"""


def write_sheet(ws, title, header_fill, headers, rows):
    ws.title = title
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    for row in rows:
        ws.append(list(row))
    ws.freeze_panes = 'A2'
    # auto width (rough)
    for c in range(1, len(headers) + 1):
        col = get_column_letter(c)
        max_len = len(str(headers[c - 1]))
        for r in range(2, min(len(rows), 200) + 2):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            s = str(v) if not isinstance(v, datetime) else v.strftime('%Y-%m-%d %H:%M')
            wlen = sum(2 if ord(ch) > 127 else 1 for ch in s)
            if wlen > max_len:
                max_len = wlen
        ws.column_dimensions[col].width = min(max_len + 2, 40)


def main():
    print(f'连数据库 ...')
    cn = pyodbc.connect(CONN)
    cur = cn.cursor()

    print(f'查汇总 ({START} ~ {END_EXCL}) ...')
    cur.execute(SUMMARY_SQL, START, END_EXCL)
    summary_cols = [c[0] for c in cur.description]
    summary_rows = cur.fetchall()
    print(f'  汇总: {len(summary_rows)} 行')

    print('查明细 ...')
    cur.execute(DETAIL_SQL, START, END_EXCL)
    detail_cols = [c[0] for c in cur.description]
    detail_rows = cur.fetchall()
    print(f'  明细: {len(detail_rows)} 行')

    cn.close()

    print(f'写 Excel: {OUT}')
    wb = Workbook()
    write_sheet(wb.active, '订单汇总',
                PatternFill('solid', fgColor='1F4E78'),
                summary_cols, summary_rows)
    write_sheet(wb.create_sheet(),  '订单明细',
                PatternFill('solid', fgColor='2E7D32'),
                detail_cols, detail_rows)
    wb.save(OUT)
    size = os.path.getsize(OUT)
    print(f'完成。文件大小: {size/1024:.1f} KB')


if __name__ == '__main__':
    main()
