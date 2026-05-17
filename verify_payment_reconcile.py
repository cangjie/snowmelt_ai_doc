#!/usr/bin/env python3
"""只读对账：按订单号汇总「支付明细」与「支付流水」，校验最终金额是否一致。

最终金额 = 支付金额 − 退款金额 − 分账金额

两个 sheet 的口径差异（来自 add_payment_detail_sheet_to_fy_xlsx.py）：
  - 支付流水：分账仅含 success=1 AND valid=1（成功分账）
  - 支付明细：分账k金额 列含全部状态（是/否/作废/空）
所以分别用「全部分账」与「仅成功分账」两种口径对比，定位差异根因。

只读，不改 DB、不改 xlsx。
用法： python3 verify_payment_reconcile.py [--xlsx path]
"""
import argparse
import os
from collections import defaultdict

import openpyxl

BASE = os.path.dirname(os.path.abspath(__file__))
DEFAULT_XLSX = os.path.join(BASE, "wanlong_rent_orders_fy_2025-05-01_2026-04-30.xlsx")
DETAIL_SHEET = "支付明细"
TX_SHEET = "支付流水"
EPS = 0.005  # 1 分以内视为相等


def r2(x):
    return round(x or 0.0, 2)


def aggregate_detail(ws):
    """支付明细 → dict[订单号] = {'pay','refund','share_all','share_ok'}"""
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    h = {name: i for i, name in enumerate(header)}
    i_code = h["订单号"]
    i_pay = h["支付金额"]
    i_refund = h["退款金额"]
    # 动态分账列：分账k金额 / 分账k成功
    share_cols = []  # (amt_idx, ok_idx)
    k = 1
    while f"分账{k}金额" in h:
        share_cols.append((h[f"分账{k}金额"], h[f"分账{k}成功"]))
        k += 1

    agg = defaultdict(lambda: {"pay": 0.0, "refund": 0.0, "share_all": 0.0, "share_ok": 0.0})
    n_rows = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        code = row[i_code]
        if code is None or str(code).strip() == "":
            continue
        code = str(code).strip()
        n_rows += 1
        a = agg[code]
        a["pay"] += float(row[i_pay] or 0)
        a["refund"] += float(row[i_refund] or 0)
        for amt_idx, ok_idx in share_cols:
            amt = row[amt_idx]
            if amt is None:
                continue
            amt = float(amt)
            a["share_all"] += amt
            if row[ok_idx] == "是":
                a["share_ok"] += amt
    return agg, n_rows, len(share_cols)


def aggregate_tx(ws):
    """支付流水 → dict[订单号] = 交易金额(含符号)合计；并按类型拆分核对。"""
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    h = {name: i for i, name in enumerate(header)}
    i_code = h["订单号"]
    i_amt = h["交易金额"]
    i_type = h["类型"]

    signed = defaultdict(float)
    by_type = defaultdict(lambda: {"支付": 0.0, "退款": 0.0, "分账": 0.0})
    n_rows = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        code = row[i_code]
        if code is None or str(code).strip() == "":
            continue
        code = str(code).strip()
        amt = float(row[i_amt] or 0)
        n_rows += 1
        signed[code] += amt
        by_type[code][row[i_type]] += amt  # 退款/分账已是负数
    return signed, by_type, n_rows


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default=DEFAULT_XLSX)
    ap.add_argument("--show", type=int, default=25, help="差异订单最多列出多少行")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.xlsx, read_only=True, data_only=True)
    for s in (DETAIL_SHEET, TX_SHEET):
        if s not in wb.sheetnames:
            raise SystemExit(f"sheet「{s}」不存在，实际: {wb.sheetnames}")

    detail, n_detail, max_share = aggregate_detail(wb[DETAIL_SHEET])
    tx_signed, tx_by_type, n_tx = aggregate_tx(wb[TX_SHEET])
    wb.close()

    print(f"支付明细：{n_detail} 行 → {len(detail)} 个订单号（动态分账列 maxShare={max_share}）")
    print(f"支付流水：{n_tx} 行 → {len(tx_signed)} 个订单号")

    all_codes = set(detail) | set(tx_signed)
    only_detail = sorted(set(detail) - set(tx_signed))
    only_tx = sorted(set(tx_signed) - set(detail))
    print(f"订单号并集 {len(all_codes)}；仅支付明细有 {len(only_detail)}；仅支付流水有 {len(only_tx)}")

    # 总额对账
    tot_pay = sum(d["pay"] for d in detail.values())
    tot_refund = sum(d["refund"] for d in detail.values())
    tot_share_all = sum(d["share_all"] for d in detail.values())
    tot_share_ok = sum(d["share_ok"] for d in detail.values())
    tot_tx_signed = sum(tx_signed.values())
    tx_pay = sum(t["支付"] for t in tx_by_type.values())
    tx_refund = sum(t["退款"] for t in tx_by_type.values())
    tx_share = sum(t["分账"] for t in tx_by_type.values())

    print("\n=== 总额对账 ===")
    print(f"支付明细  Σ支付金额        = {tot_pay:>15,.2f}")
    print(f"支付流水  Σ支付(正)        = {tx_pay:>15,.2f}   差 {tot_pay - tx_pay:+.2f}")
    print(f"支付明细  Σ退款金额        = {tot_refund:>15,.2f}")
    print(f"支付流水  Σ退款(abs)       = {-tx_refund:>15,.2f}   差 {tot_refund - (-tx_refund):+.2f}")
    print(f"支付明细  Σ分账(全部状态)  = {tot_share_all:>15,.2f}")
    print(f"支付明细  Σ分账(仅成功=是) = {tot_share_ok:>15,.2f}")
    print(f"支付流水  Σ分账(abs,成功)  = {-tx_share:>15,.2f}   vs 仅成功差 {tot_share_ok - (-tx_share):+.2f}")

    detail_final_all = tot_pay - tot_refund - tot_share_all
    detail_final_ok = tot_pay - tot_refund - tot_share_ok
    print(f"\n支付明细 最终金额(分账=全部) = {detail_final_all:>15,.2f}")
    print(f"支付明细 最终金额(分账=仅成功)= {detail_final_ok:>15,.2f}")
    print(f"支付流水 Σ交易金额(含符号)   = {tot_tx_signed:>15,.2f}")
    print(f"  vs 全部口径 差 {detail_final_all - tot_tx_signed:+.2f}")
    print(f"  vs 仅成功口径 差 {detail_final_ok - tot_tx_signed:+.2f}")

    # 逐订单对账（用「仅成功分账」口径，与支付流水同口径）
    mism_ok = []
    mism_all = []
    for code in sorted(all_codes):
        d = detail.get(code, {"pay": 0, "refund": 0, "share_all": 0, "share_ok": 0})
        fin_ok = r2(d["pay"] - d["refund"] - d["share_ok"])
        fin_all = r2(d["pay"] - d["refund"] - d["share_all"])
        tx = r2(tx_signed.get(code, 0.0))
        if abs(fin_ok - tx) > EPS:
            mism_ok.append((code, fin_ok, tx, fin_ok - tx, d))
        if abs(fin_all - tx) > EPS:
            mism_all.append((code, fin_all, tx, fin_all - tx, d))

    print("\n=== 逐订单对账 ===")
    print(f"口径A 支付明细(分账=仅成功) vs 支付流水：不一致 {len(mism_ok)} 单")
    print(f"口径B 支付明细(分账=全部)   vs 支付流水：不一致 {len(mism_all)} 单")

    if mism_ok:
        print(f"\n--- 口径A 不一致明细（前 {args.show}）订单号 | 支付明细最终 | 支付流水 | 差 ---")
        for code, fin, tx, diff, d in sorted(mism_ok, key=lambda x: -abs(x[3]))[: args.show]:
            print(f"  {code:<24} {fin:>12,.2f} {tx:>12,.2f} {diff:>+10,.2f}"
                  f"  [支付{d['pay']:.2f} 退款{d['refund']:.2f} 分账成功{d['share_ok']:.2f}]")
        print(f"  口径A 差额合计: {sum(m[3] for m in mism_ok):+,.2f}")

    if mism_all:
        diff_only_all = [m for m in mism_all if m[0] not in {x[0] for x in mism_ok}]
        print(f"\n--- 口径B 额外多出的不一致（含失败/作废分账，不在口径A 内）{len(diff_only_all)} 单 ---")
        for code, fin, tx, diff, d in sorted(diff_only_all, key=lambda x: -abs(x[3]))[: args.show]:
            print(f"  {code:<24} {fin:>12,.2f} {tx:>12,.2f} {diff:>+10,.2f}"
                  f"  [分账全部{d['share_all']:.2f} 分账成功{d['share_ok']:.2f}]")

    print("\n结论：")
    if not mism_ok:
        print("  ✓ 用「仅成功分账」口径，两 sheet 按订单号最终金额完全一致。")
    else:
        print(f"  ✗ 即使用「仅成功分账」口径仍有 {len(mism_ok)} 单不一致，需排查（见上）。")
    if mism_all and not mism_ok:
        print("  ⚠ 用「全部分账」口径会出现差异，根因=支付明细含失败/作废分账，支付流水只收成功分账（设计如此）。")


if __name__ == "__main__":
    main()
